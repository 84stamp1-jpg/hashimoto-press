#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
橋本工業 生産計画自動生成スクリプト v2.0
使い方: python generate_plan.py [--date YYYY-MM-DD] [--workers N] [--slack]
"""
import argparse, math, os, sys
from copy import deepcopy
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ─── パス設定 ────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
INPUT_DIR   = BASE_DIR / "input"
OUTPUT_DIR  = BASE_DIR / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
ORDER_FILE  = INPUT_DIR / "作業予定一覧.xlsx"
MASTER_FILE = INPUT_DIR / "製品マスタ.xlsx"

# ─── 定数 ────────────────────────────────────────────
WS      = 8 * 60 + 10   # 490 始業 8:10
WE      = 17 * 60       # 1020 終業 17:00
SLOT    = 15            # 15分刻み
NS      = (WE - WS) // SLOT   # 35スロット
BREAKS  = [(600, 610), (730, 770), (890, 900)]
WORK_MIN = (WE - WS) - sum(e - s for s, e in BREAKS)   # 470分
DOW_JP  = ["月", "火", "水", "木", "金", "土", "日"]

def is_break(t):
    for s, e in BREAKS:
        if s <= t < e: return True
    return False

def adv(start, mins):
    t, r = start, int(mins)
    while r > 0 and t < WE:
        if not is_break(t): r -= 1
        t += 1
    return t

def work_in(s, e):
    return sum(1 for t in range(max(WS, s), min(WE, e)) if not is_break(t))

def nw(t):
    while t < WE and is_break(t): t += 1
    return t

def hhmm(v):
    h, m = divmod(int(v), 60)
    return f"{h:02d}:{m:02d}"

def sl(v):
    return (int(v) - WS) // SLOT

# ─── データ読み込み ───────────────────────────────────
def load_master():
    """製品マスタを読み込む（3行ヘッダー構造）"""
    raw = pd.read_excel(MASTER_FILE, sheet_name="マスタ本体", header=None)
    cols = [str(v).strip() if pd.notna(v) else f"_c{i}"
            for i, v in enumerate(raw.iloc[2])]
    df = raw.iloc[5:].copy()
    df.columns = cols
    df = df.reset_index(drop=True)
    df = df[df["品番"].apply(
        lambda x: pd.notna(x) and str(x).strip() not in ["nan", ""]
    )].copy()
    df["品番"] = df["品番"].astype(str).str.strip()
    for col in ["段取時間(分)", "段取人員(名)", "取数(個)",
                "標準時間1(個/h)", "標準作業員(名)", "最小作業員(名)"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df

def load_patterns():
    """加工パターンテーブルを読み込む（3行ヘッダー構造）"""
    raw = pd.read_excel(MASTER_FILE, sheet_name="加工パターンテーブル", header=None)
    cols = [str(v).strip() if pd.notna(v) else f"_p{i}"
            for i, v in enumerate(raw.iloc[2])]
    df = raw.iloc[3:].copy()
    df.columns = cols
    df = df.reset_index(drop=True)
    df = df[df["品番"].apply(
        lambda x: pd.notna(x) and str(x).strip() not in ["nan", ""]
        and "スケジューラ" not in str(x) and "段取の扱い" not in str(x)
    )].copy()
    df["品番"] = df["品番"].astype(str).str.strip()
    for col in ["作業員\n（名）", "段取時間\n（分）", "段取人員\n（名）",
                "出来高\n（個/h）", "取数\n（個）", "優先度"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df[df["出来高\n（個/h）"].notna()].copy()

def load_orders():
    """作業予定一覧を読み込む（プレス部門のみ）"""
    df = pd.read_excel(ORDER_FILE, header=0)
    df = df[df["製造番号"].notna()].copy()
    df["品番"] = df["品番"].astype(str).str.strip()
    df["作業票納期"] = pd.to_datetime(df["作業票納期"])
    df["due"] = df["作業票納期"].dt.date
    return df[df["部門"] == "プレス"].copy()

def build_jobs(master, patterns, orders):
    """ジョブリストを生成"""
    def norm(s):
        return str(s).replace("　", "").replace(" ", "").strip() if pd.notna(s) else ""

    orders["工程_norm"] = orders["工程"].apply(norm)
    master["工程_norm"] = master["工程(ライン)"].apply(norm)
    master_parts = set(master["品番"].tolist())

    mc = ["品番", "工程_norm", "段取時間(分)", "段取人員(名)", "取数(個)",
          "標準時間1(個/h)", "標準作業員(名)", "最小作業員(名)",
          "人員削減タイプ", "設備1", "設備2", "設備3", "設備4", "代替設備リスト"]
    avail = [c for c in mc if c in master.columns]
    merged = orders.merge(master[avail], on=["品番", "工程_norm"], how="left")

    mf = master.drop_duplicates("品番", keep="first")
    for i in merged[merged.get("標準時間1(個/h)", pd.Series()).isna()
                    if "標準時間1(個/h)" in merged.columns else merged.index].index:
        pn = merged.at[i, "品番"]
        row_m = mf[mf["品番"] == pn]
        if not row_m.empty:
            for c in avail[2:]:
                if c in row_m.columns:
                    merged.at[i, c] = row_m.iloc[0][c]

    jobs = []
    for _, r in merged.iterrows():
        pn   = r["品番"]
        std1 = float(r["標準時間1(個/h)"]) if pd.notna(r.get("標準時間1(個/h)")) else 0
        tori = float(r["取数(個)"])         if pd.notna(r.get("取数(個)"))         else 1

        def fv(nc, oc, default):
            try:
                v = float(r.get(nc, None))
                if v > 0: return v
            except: pass
            try:
                v2 = float(r.get(oc, None))
                if pd.notna(v2): return v2
            except: pass
            return default

        setup_m = fv("段取時間(分)", "段取(M)", 30)
        work_w  = fv("標準作業員(名)", "投入人数", 1)
        setup_w = float(r["段取人員(名)"]) if pd.notna(r.get("段取人員(名)")) else work_w
        min_w   = float(r["最小作業員(名)"]) if pd.notna(r.get("最小作業員(名)")) else work_w
        qty     = int(r["加工指示数"]) if pd.notna(r["加工指示数"]) else 0
        due     = r["due"]

        devs = [str(r.get(f"設備{i}", "")).strip() for i in range(1, 5)
                if pd.notna(r.get(f"設備{i}")) and
                str(r.get(f"設備{i}", "")).strip() not in ["nan", ""]]

        if std1 > 0:
            ppm = std1 * tori / 60.0
        elif pd.notna(r.get("加工(M)")) and float(r["加工(M)"]) > 0:
            ppm = 1.0 / float(r["加工(M)"]) / 60.0
        else:
            ppm = 1 / 60.0

        # 設備なし → 仮想設備
        if not devs:
            ln = str(r.get("工程", "")).replace("プレス", "").replace("　", "").strip()
            devs = [f"{ln}_MISC"]

        # パターン一覧
        pats = patterns[patterns["品番"] == pn].sort_values("優先度").to_dict("records")

        jobs.append(dict(
            mfg=r["製造番号"], jun=int(r["工順"]),
            line=norm(r["工程"]), part=pn, name=str(r["品名"]).strip()[:32],
            devs=devs, dev="+".join(devs),
            std1h=std1 * tori, ppm=ppm,
            setup_m=setup_m, setup_w=int(setup_w),
            work_w=int(work_w), min_w=int(min_w),
            qty=qty, remain=qty, due=due,
            not_in_master=(pn not in master_parts),
            patterns=pats,
        ))
    return jobs

# ─── スケジューラ ─────────────────────────────────────
def get_pattern(j):
    """パターンテーブルから最適パターンを返す"""
    if j["patterns"]:
        pat = j["patterns"][0]
        p_devs = [str(pat.get(f"使用設備{i}", "")).strip() for i in range(1, 6)
                  if pd.notna(pat.get(f"使用設備{i}")) and
                  str(pat.get(f"使用設備{i}", "")).strip() not in ["nan", ""]]
        if not p_devs: p_devs = j["devs"]
        return {
            "devs":     p_devs,
            "work_w":   int(pat["作業員\n（名）"]),
            "setup_w":  int(pat["段取人員\n（名）"]),
            "setup_m":  int(pat["段取時間\n（分）"]),
            "ppm":      float(pat["出来高\n（個/h）"]) / 60,
            "pat_name": pat["パターン名"],
        }
    return {
        "devs":     j["devs"], "work_w": j["work_w"], "setup_w": j["setup_w"],
        "setup_m":  j["setup_m"], "ppm": j["ppm"], "pat_name": "標準",
    }

def schedule_day(day, jobs_pool, remain_map, max_workers):
    wtl = [0] * (WE - WS)
    dev_busy = {}

    def can(s, e, w):
        for t in range(max(WS, s), min(WE, e)):
            if not is_break(t) and wtl[t - WS] + w > max_workers:
                return False
        return True

    def add(s, e, w):
        for t in range(max(WS, s), min(WE, e)):
            if not is_break(t): wtl[t - WS] += w

    def earliest(w, after=WS):
        t = nw(after)
        while t < WE:
            if not is_break(t) and wtl[t - WS] + w <= max_workers:
                return t
            t += 1
        return WE

    sched = []
    for j in sorted([j for j in jobs_pool if remain_map.get((j["mfg"], j["jun"]), 0) > 0],
                    key=lambda x: ((x["due"] - day).days, x["jun"])):
        remain = remain_map.get((j["mfg"], j["jun"]), 0)
        if remain <= 0: continue

        cp = get_pattern(j)
        devs, ww, sw, sm, ppm = (
            cp["devs"], cp["work_w"], cp["setup_w"], cp["setup_m"], cp["ppm"]
        )

        # 設備空き
        dev_start = max((dev_busy.get(d, WS) for d in devs), default=WS)
        dev_start = nw(max(dev_start, WS))
        if dev_start >= WE: continue

        # 段取開始（設備空き & 段取員が入れる）
        ss = dev_start
        if not can(ss, adv(ss, sm + 1), sw):
            ss = max(earliest(sw, ss), dev_start)
            ss = nw(ss)
        dev_start2 = max((dev_busy.get(d, WS) for d in devs), default=WS)
        if ss < dev_start2: ss = nw(dev_start2)
        if ss >= WE: continue
        if not can(ss, adv(ss, sm + 1), sw): continue

        se = adv(ss, sm)
        ws_ = nw(se)
        if ws_ >= WE: continue

        # 加工時間を仮計算して全体チェック（★重要：1分後ではなく全体）
        tq = min(remain, int(work_in(ws_, WE) * ppm))
        if tq <= 0: continue
        wm = math.ceil(tq / ppm)
        we_trial = min(adv(ws_, wm), WE)

        if not can(ws_, we_trial, ww):
            # 入れる時刻を全体チェックで探す
            found = False
            for try_ws in range(ws_, WE):
                if is_break(try_ws): continue
                we_t = min(adv(try_ws, wm), WE)
                if can(try_ws, we_t, ww):
                    ws_ = try_ws; we_trial = we_t; found = True; break
            if not found: continue

        we_ = we_trial
        aq = min(remain, int(work_in(ws_, we_) * ppm))
        if aq <= 0: continue

        add(ss, se, sw)      # 段取中: sw名
        add(ws_, we_, ww)    # 加工中: ww名
        for d in devs: dev_busy[d] = we_

        sched.append({
            **j, "remain": remain, "aq": aq,
            "ss": ss, "se": se, "ws": ws_, "we": we_,
            "is_over": (j["due"] < day), "is_today": (j["due"] == day),
            "day": day, "chosen_pat": cp,
            "devs_used": devs, "work_w_used": ww,
            "setup_w_used": sw, "setup_m_used": sm,
        })
    return sched, wtl

# ─── Excel出力 ────────────────────────────────────────
PAL = {
    "title":"1F4E79","hdr":"2E75B6","hdr2":"4472C4","wt":"FFFFFF",
    "br":"C8C8C8","br2":"E4E4E4","setup":"FFC000","setup_lt":"FFE780",
    "PA":"4472C4","PB":"ED7D31","PC":"70AD47","PD":"00B0F0",
    "gA":"D6E4F7","gB":"FDE9D9","gC":"E2EFDA","gD":"DDEBF7",
    "rbg":"F8FBFF","over":"FFE0E0","today":"E8F5E9","miss_bg":"FFF8F0",
    "w13":"00B050","w10":"FFEB9C","w7":"FDE9D9","idle":"FFF0CC",
    "dark":"1A1A1A","gray":"888888","w0":"F3F3F3",
}
GK = {"プレスA":"A","プレスB":"B","プレスC":"C","プレスD":"D"}
BAR = {"プレスA":"PA","プレスB":"PB","プレスC":"PC","プレスD":"PD"}
LO = ["プレスA","プレスB","プレスC","プレスD"]

def F(h): return PatternFill("solid",start_color=h,end_color=h)
def FN(b=False,c="000000",s=8): return Font(name="Meiryo UI",bold=b,color=c,size=s)
def AL(h="center",v="center",w=False): return Alignment(horizontal=h,vertical=v,wrap_text=w)
def dk(h,f=.72):
    r,g,b=int(h[:2],16),int(h[2:4],16),int(h[4:],16)
    return f"{max(0,int(r*f)):02X}{max(0,int(g*f)):02X}{max(0,int(b*f)):02X}"
def lt(h,f=.35):
    r,g,b=int(h[:2],16),int(h[2:4],16),int(h[4:],16)
    return f"{min(255,int(r+(255-r)*f)):02X}{min(255,int(g+(255-g)*f)):02X}{min(255,int(b+(255-b)*f)):02X}"
_t=Side(style="thin",color="BFBFBF"); _s=Side(style="thin",color="E0E0E0")
BD=Border(left=_t,right=_t,top=_t,bottom=_t)
SBD=Border(left=_s,right=_s,top=_s,bottom=_s)

def build_sheet(wb, day, sched, wtl, max_workers):
    dow=DOW_JP[day.weekday()]
    ws=wb.create_sheet(f"{day.month:02d}-{day.day:02d}({dow})")
    ws.sheet_view.showGridLines=False; ws.sheet_view.zoomScale=88
    FIX=6; TL=FIX+1; end_c=get_column_letter(TL+NS-1)
    for i,w in enumerate([9,20,30,8,6,9]):
        ws.column_dimensions[get_column_letter(i+1)].width=w
    for i in range(NS):
        am=WS+i*SLOT
        ws.column_dimensions[get_column_letter(TL+i)].width=2.8 if not is_break(am) else 1.4
    ws.row_dimensions[1].height=26; ws.row_dimensions[2].height=14; ws.row_dimensions[3].height=12
    peak=max(wtl) if wtl else 0
    wku=sum(wtl[t-WS] for t in range(WS,WE) if not is_break(t))
    util=wku/(max_workers*WORK_MIN)*100
    n_miss=sum(1 for j in sched if j.get("not_in_master"))
    ws.merge_cells(f"A1:{end_c}1")
    c=ws["A1"]
    c.value=(f"生産計画　{day.year}/{day.month}/{day.day}（{dow}）　"
             f"出勤{max_workers}名  計画{len(sched)}件  ピーク{peak}名  活用率{util:.0f}%"
             +(f"  ★マスタ未登録{n_miss}件" if n_miss else "")
             +"　　休憩: 10:00-10:10 / 12:10-12:50 / 14:50-15:00")
    c.font=FN(True,PAL["wt"],10); c.fill=F(PAL["title"]); c.alignment=AL()

    def mhdr(r1,c1,r2,c2,v,col):
        if r1!=r2 or c1!=c2:
            ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
        cell=ws.cell(r1,c1,v)
        cell.font=FN(True,PAL["wt"],8); cell.fill=F(col)
        cell.alignment=AL(w=True); cell.border=BD

    mhdr(2,1,3,1,"工程",PAL["hdr"]); mhdr(2,2,3,2,"設備・機械",PAL["hdr"])
    mhdr(2,3,2,3,"品　番",PAL["hdr2"]); mhdr(3,3,3,3,"品　名",PAL["hdr"])
    mhdr(2,4,2,4,"指示数",PAL["hdr2"]); mhdr(3,4,3,4,"本日生産",PAL["hdr"])
    mhdr(2,5,2,5,"段取員",PAL["hdr2"]); mhdr(3,5,3,5,"作業員",PAL["hdr"])
    mhdr(2,6,3,6,"時　間",PAL["hdr"])

    done=set()
    for i in range(NS):
        am=WS+i*SLOT; h_,m_=divmod(am,60); col_=TL+i; ib=is_break(am)
        if m_==0 and h_ not in done:
            ws.merge_cells(start_row=2,start_column=col_,end_row=2,end_column=min(col_+3,TL+NS-1))
            c=ws.cell(2,col_,f"{h_:02d}:00"); c.font=FN(True,PAL["wt"],8)
            c.fill=F(PAL["br"] if ib else PAL["title"]); c.alignment=AL(); c.border=BD; done.add(h_)
        c2=ws.cell(3,col_,f"{m_:02d}" if m_%30==0 else ""); c2.font=FN(c=PAL["wt"],s=6)
        c2.fill=F(PAL["br2"] if ib else PAL["hdr"]); c2.alignment=AL(); c2.border=BD

    if not sched:
        ws.row_dimensions[4].height=28; ws.merge_cells(f"A4:{end_c}4")
        c=ws.cell(4,1,"この日の生産計画はありません")
        c.font=FN(True,PAL["gray"],10); c.fill=F(PAL["w0"]); c.alignment=AL()
        ws.freeze_panes=f"{get_column_letter(TL)}4"; return ws

    row=4
    for line in LO:
        grp=[j for j in sched if j["line"]==line]
        if not grp: continue
        gk=GK[line]; gbg=PAL[f"g{gk}"]; bar=PAL[BAR[line]]
        ws.row_dimensions[row].height=13
        ws.merge_cells(f"A{row}:{end_c}{row}")
        c=ws.cell(row,1,f"■  {line}　（{len(grp)}件　段取員:{sum(j['setup_w_used'] for j in grp)}名　加工員:{sum(j['work_w_used'] for j in grp)}名）")
        c.font=FN(True,PAL["dark"],8); c.fill=F(dk(gbg)); c.alignment=AL("left")
        for ci in range(1,TL+NS): ws.cell(row,ci).fill=F(dk(gbg))
        row+=1
        for j in sorted(grp,key=lambda x:x["ss"]):
            is_ov=j["is_over"]; is_td=j["is_today"]; is_miss=j.get("not_in_master",False)
            rbg_u=PAL["over"] if is_ov else (PAL["today"] if is_td else (PAL["miss_bg"] if is_miss else lt(gbg,0.2)))
            rbg_d=PAL["over"] if is_ov else (PAL["today"] if is_td else (PAL["miss_bg"] if is_miss else lt(gbg,0.45)))
            qty_orig=j.get("qty",j["remain"]+j["aq"])
            pat_str=j.get("chosen_pat",{}).get("pat_name","")
            pat_d=f" [{pat_str}]" if pat_str and pat_str!="標準" else ""
            ws.row_dimensions[row].height=17; ws.row_dimensions[row+1].height=13
            def mc(r1,c1,r2,c2): ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
            mc(row,1,row+1,1); c=ws.cell(row,1,line); c.font=FN(s=7); c.fill=F(rbg_u); c.alignment=AL(); c.border=BD
            mc(row,2,row+1,2)
            dev_d="+".join(j["devs_used"]).replace("_MISC","").replace("B_","").replace("C_","")
            c=ws.cell(row,2,dev_d); c.font=FN(b=True,s=8); c.fill=F(rbg_u); c.alignment=AL("left"); c.border=BD
            miss_s="★" if is_miss else ""
            fc_pn="CC0000" if is_ov else ("CC4400" if is_miss else "1A3A6B")
            c=ws.cell(row,3,f"{miss_s}{j['part']}{pat_d}"); c.font=FN(b=True,s=8,c=fc_pn); c.fill=F(rbg_u); c.alignment=AL("left"); c.border=BD
            c=ws.cell(row+1,3,j["name"]); c.font=FN(s=7,c="445566"); c.fill=F(rbg_d); c.alignment=AL("left"); c.border=BD
            c=ws.cell(row,4,qty_orig); c.font=FN(s=8); c.fill=F(rbg_u); c.alignment=AL("right"); c.border=BD; c.number_format="#,##0"
            fc_aq="00AA00" if j["aq"]>=qty_orig else PAL["dark"]
            c=ws.cell(row+1,4,j["aq"]); c.font=FN(b=True,s=8,c=fc_aq); c.fill=F(rbg_d); c.alignment=AL("right"); c.border=BD; c.number_format="#,##0"
            c=ws.cell(row,5,f"{j['setup_w_used']}名"); c.font=FN(s=8,c=PAL["gray"]); c.fill=F(rbg_u); c.alignment=AL(); c.border=BD
            c=ws.cell(row+1,5,f"{j['work_w_used']}名"); c.font=FN(b=True,s=8); c.fill=F(rbg_d); c.alignment=AL(); c.border=BD
            mc(row,6,row+1,6); c=ws.cell(row,6,f"{hhmm(j['ss'])}〜{hhmm(j['we'])}"); c.font=FN(s=7); c.fill=F(rbg_u); c.alignment=AL(); c.border=BD
            for i in range(NS):
                am=WS+i*SLOT; ib=is_break(am)
                for r2 in [row,row+1]:
                    ws.cell(r2,TL+i).fill=F(PAL["br2"] if ib else PAL["rbg"]); ws.cell(r2,TL+i).border=SBD
            ss_s=sl(j["ss"]); se_s=sl(j["se"]); ws_s=sl(j["ws"]); we_s=sl(j["we"])
            for i in range(ss_s,min(se_s,NS)):
                if not is_break(WS+i*SLOT):
                    ws.cell(row,TL+i).fill=F(PAL["setup"]); ws.cell(row,TL+i).border=BD
                    ws.cell(row+1,TL+i).fill=F(PAL["setup_lt"]); ws.cell(row+1,TL+i).border=BD
            for i in range(ws_s,min(we_s,NS)):
                if not is_break(WS+i*SLOT):
                    ws.cell(row,TL+i).fill=F(bar); ws.cell(row,TL+i).border=BD
                    ws.cell(row+1,TL+i).fill=F(lt(bar,0.4)); ws.cell(row+1,TL+i).border=BD
            if we_s-1<NS:
                c=ws.cell(row,TL+min(we_s-1,NS-1)); c.value="◀"; c.font=FN(True,PAL["wt"],5); c.alignment=AL(h="right")
            row+=2

    for brs,bre in BREAKS:
        for si in range(sl(brs),min(sl(bre),NS)):
            for r2 in range(2,row+3): ws.cell(r2,TL+si).fill=F(PAL["br"])

    row+=1; ws.row_dimensions[row].height=12
    ws.merge_cells(f"A{row}:{get_column_letter(FIX)}{row}")
    c=ws.cell(row,1,"▼ 時間帯別 稼働人員（段取員・加工員）  橙色=手待ち"); c.font=FN(True,PAL["wt"],8); c.fill=F(PAL["title"]); c.alignment=AL("left")
    for ci in range(FIX+1,TL+NS): ws.cell(row,ci).fill=F(PAL["title"])
    row+=1; ws.row_dimensions[row].height=16
    ws.merge_cells(f"A{row}:{get_column_letter(FIX)}{row}")
    c=ws.cell(row,1,"稼働人員数"); c.font=FN(True,s=8); c.fill=F("DDEBF7"); c.alignment=AL("left")
    for ci in range(2,FIX+1): ws.cell(row,ci).fill=F("DDEBF7")
    for i in range(NS):
        am=WS+i*SLOT; ib=is_break(am); w=wtl[i*SLOT] if i*SLOT<len(wtl) else 0
        c=ws.cell(row,TL+i)
        if ib: c.value="休"; c.font=FN(s=5,c=PAL["gray"]); c.fill=F(PAL["br"]); c.alignment=AL(); c.border=BD
        else:
            c.value=int(w) if w>0 else ""
            if w>=max_workers: bg,fc=PAL["w13"],PAL["wt"]
            elif w>=10:        bg,fc=PAL["w10"],PAL["dark"]
            elif w>=7:         bg,fc=PAL["w7"],PAL["dark"]
            elif w>0:          bg,fc="FDE9D9",PAL["dark"]
            else:              bg,fc=PAL["idle"],"CC8800"
            c.fill=F(bg); c.font=FN(True,s=7,c=fc); c.alignment=AL(); c.border=BD

    miss_jobs=[j for j in sched if j.get("not_in_master")]
    if miss_jobs:
        row+=2; ws.row_dimensions[row].height=13
        ws.merge_cells(f"A{row}:{end_c}{row}")
        c=ws.cell(row,1,f"★ マスタ未登録品番（{len(miss_jobs)}件）— 作業予定一覧の加工時間から出来高を計算（精度低）")
        c.font=FN(True,"CC4400",9); c.fill=F("FFF8F0"); c.alignment=AL("left")
        for j in miss_jobs:
            row+=1; ws.row_dimensions[row].height=14
            ws.merge_cells(f"A{row}:{end_c}{row}")
            c=ws.cell(row,1,f"  • {j['part']}  ({j['name']})  製造番号:{j['mfg']}  納期:{j['due']}  → マスタへの登録をご確認ください")
            c.font=FN(s=8,c="884400"); c.fill=F("FFF8F0"); c.alignment=AL("left")

    row+=2; ws.row_dimensions[row].height=13
    leg=[(PAL["setup"],"段取"),(PAL["PA"],"プレスA"),(PAL["PB"],"プレスB"),
         (PAL["PC"],"プレスC"),(PAL["PD"],"プレスD"),(PAL["br"],"休憩"),
         (PAL["over"],"納期超過"),(PAL["today"],"当日納期"),
         (PAL["w13"],f"{max_workers}名フル"),(PAL["w10"],"10〜12名"),(PAL["idle"],"手待ち")]
    ws.cell(row,1,"凡例").font=FN(True,s=8)
    for idx,(col,lbl) in enumerate(leg):
        cc=2+idx*2
        if cc+1>TL+NS: break
        c=ws.cell(row,cc,"　"); c.fill=F(col); c.border=BD
        ws.cell(row,cc+1,lbl).font=FN(s=7)
    ws.freeze_panes=f"{get_column_letter(TL)}4"
    return ws

# ─── メイン ──────────────────────────────────────────
def main():
    parser=argparse.ArgumentParser(description="橋本工業 生産計画自動生成 v2.0")
    parser.add_argument("--date",    default=None)
    parser.add_argument("--workers", type=int, default=13)
    parser.add_argument("--slack",   action="store_true")
    parser.add_argument("--week",    action="store_true", help="1週間分（省略時も1週間）")
    args=parser.parse_args()

    target=date.fromisoformat(args.date) if args.date else date.today()
    while target.weekday()>=5: target+=timedelta(1)

    print(f"[{datetime.now():%H:%M:%S}] 生産計画生成開始  対象:{target}  出勤:{args.workers}名", flush=True)

    if not ORDER_FILE.exists():  sys.exit(f"エラー: {ORDER_FILE} が見つかりません")
    if not MASTER_FILE.exists(): sys.exit(f"エラー: {MASTER_FILE} が見つかりません")

    master   = load_master()
    patterns = load_patterns()
    orders   = load_orders()
    jobs     = build_jobs(master, patterns, orders)

    miss=[j for j in jobs if j["not_in_master"]]
    if miss:
        print(f"  [警告] マスタ未登録品番: {len(miss)}件")
        for j in miss: print(f"    {j['part']} ({j['name']}) due:{j['due']}")

    # 稼働日リスト（5日間）
    work_days=[]
    d=target
    while len(work_days)<5:
        if d.weekday()<5: work_days.append(d)
        d+=timedelta(1)

    pool     = deepcopy(jobs)
    remain   = {(j["mfg"],j["jun"]): j["remain"] for j in pool}
    week_plan= {}

    for day in work_days:
        active=[j for j in pool if remain.get((j["mfg"],j["jun"]),0)>0]
        sched,wtl=schedule_day(day,active,remain,args.workers)
        for s in sched:
            remain[(s["mfg"],s["jun"])]=max(0,remain.get((s["mfg"],s["jun"]),0)-s["aq"])
        week_plan[day]=(sched,wtl)
        peak=max(wtl) if wtl else 0
        wku=sum(wtl[t-WS] for t in range(WS,WE) if not is_break(t))
        print(f"  {day}({DOW_JP[day.weekday()]}) {len(sched)}件 ピーク:{peak}名 "
              f"活用:{wku/(args.workers*WORK_MIN)*100:.1f}%")

    # Excel生成
    wb=Workbook(); wb.remove(wb.active)
    for day in work_days:
        sched,wtl=week_plan[day]
        build_sheet(wb,day,sched,wtl,args.workers)

    fname=OUTPUT_DIR/f"生産計画_{target:%Y%m%d}.xlsx"
    wb.save(fname)
    print(f"  完了: {fname}")

    # Slack投稿
    if args.slack:
        token  =os.environ.get("SLACK_BOT_TOKEN","")
        channel=os.environ.get("SLACK_CHANNEL","C06DBE9536Y")
        if not token:
            print("  [警告] SLACK_BOT_TOKEN 未設定。Slack投稿をスキップ。")
        else:
            try:
                import requests
                dow=DOW_JP[target.weekday()]
                total=sum(len(s) for s,_ in week_plan.values())
                text=(f":factory: *{target:%Y/%m/%d}（{dow}）生産計画*\n"
                      f"計画{total}件　出勤{args.workers}名")
                r=requests.post("https://slack.com/api/chat.postMessage",
                    json={"channel":channel,"text":text,"mrkdwn":True},
                    headers={"Authorization":f"Bearer {token}"})
                ts=r.json().get("ts")
                with open(fname,"rb") as f:
                    requests.post("https://slack.com/api/files.upload",
                        headers={"Authorization":f"Bearer {token}"},
                        data={"channels":channel,"thread_ts":ts or "",
                              "filename":fname.name,"title":fname.name},
                        files={"file":f})
                print(f"  Slack投稿完了 -> #{channel}")
            except Exception as e:
                print(f"  [警告] Slack投稿失敗: {e}")

    print(f"[{datetime.now():%H:%M:%S}] 完了")

if __name__=="__main__":
    main()
