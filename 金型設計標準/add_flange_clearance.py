# -*- coding: utf-8 -*-
"""金型設計クリアランス基準書：「バーリング」シートを
「フランジ成形」シート（凹円弧／凸円弧の区分ベース）へ再構成する。
※ 縁曲げ（旧称フランジダウン）とバーリングを1枚に統合したもの。

背景
  バーリングと縁曲げは同じ「板の縁を曲げ起こす成形」で、違いは曲げ線が
  凹円弧（伸びフランジ）か凸円弧（縮みフランジ）かにある。
  さらに片側クリアランスの基準となる板厚は素材板厚 t0 ではなく、
  その部位の実板厚 ta（ドロー後は t0 より薄い）である。
  この2点を反映して1枚の基準に統合する。

  実行すると 金型設計クリアランス基準書.xlsx の「バーリング」シートを
  同じ位置の「フランジ成形」シートで置き換える（旧内容は下穴逆算・注記とも継承）。
  実行前に .bak_フランジ追加前_<日付>.xlsx を自動で取る。

使い方: python add_flange_clearance.py
"""
import os
import shutil
from copy import copy
from datetime import date

import openpyxl

import build_hms_docs as B   # パスとシート名の定義を再利用（二重定義しない）

BASE = B.BASE
XLSX = B.CLEAR_XLSX
KYOTSU = os.path.join(BASE, B.EDITIONS['共通'][0])

OLD_SHEET = 'バーリング'
NEW_SHEET = 'フランジ成形'

# ---------------------------------------------------------------- シート内容
# ('title'|'sub'|'head'|'note', text) / ('table', [行...]) / ('calc', [行...])
CONTENT = [
    ('title', 'フランジ成形（バーリング／縁曲げ）クリアランス基準'),
    ('sub', 'バーリングも縁曲げも、板の縁を曲げ起こす成形である。違いは曲げ線が'
            '凹円弧か凸円弧かにある。片側クリアランス c は素材板厚 t0 ではなく、'
            'その部位の実板厚 ta を基準に決める。'),

    ('head', '■ 1. 基準板厚 ta の決め方（最重要）'),
    ('sub', 'c ＝ 係数 × ta。ta はその部位の実際の板厚で、素材板厚 t0 とは限らない。'
            'ドロー・張出しで引けている部位に成形する場合、ta は t0 より小さくなる。'),
    ('table', [
        ['部位の履歴', 'ta の目安', '確認方法'],
        ['平板から直接（前工程で成形なし）', 'ta ＝ t0', 'ミルシート＋実測'],
        ['ドロー・張出しで引けている部位', 'ta ＝ 0.70〜0.90 × t0',
         'CAEの板厚分布／トライ品を切断して実測'],
        ['絞り縁など肉が寄っている部位', 'ta ＝ 1.00〜1.20 × t0', '同上'],
        ['曲げ外側のみ（直線フランジ）', 'ta ≒ t0', '—'],
    ]),
    ('note', '※ ドロー後のバーリングを t0 基準で組むと実際にはクリアランス過大となり、'
             'フランジが低く厚く出る。必ず ta で計算する。'),
    ('note', '※ ta が不明なときは暫定 0.85×t0 で設計し、初回トライで実測して型を追い込む。'
             '実測板厚は設計値管理表（C5-1）に必ず残す。'),
    ('note', '※ 素材板厚は公差の＋側で来ることがある。詰める側（凹・穴）は ta の下限、'
             '逃がす側（凸・直線）は ta の上限でも成立するか確認する。'),
    ('note', '※ フェライト系ステン（436L/429M）は r値が高く（1.4〜1.6 対 SUS304 の 0.9〜1.0）'
             '板厚が減りにくい。ドロー後の ta は 0.70〜0.90×t0 の上側を採ってよい。'),

    ('head', '■ 2. 凹・凸の判別'),
    ('table', [
        ['区分', '曲げ線と代表例', '縁の周長', '材料の挙動', '出やすい不具合'],
        ['凹円弧（伸びフランジ）',
         'フランジが円弧の内側。バーリング（穴フランジ）、切欠き・内Rの縁曲げ',
         '伸びる', '引張。縁が薄くなる（元の約70%まで）', '縁割れ・花咲き'],
        ['直線', '直線。一般の曲げ、直線部の縁曲げ',
         '変わらない', '板厚ほぼ変化なし', 'スプリングバック'],
        ['凸円弧（縮みフランジ）',
         'フランジが円弧の外側。外周の縁曲げ、絞り縁',
         '縮む', '圧縮。肉が寄って厚くなる', 'しわ・肉余り'],
    ]),
    ('note', '※ 判別は「フランジ縁の周長が曲げ線の周長より長くなるか短くなるか」で決める。'
             '長くなる＝凹（伸び）、短くなる＝凸（縮み）。'),
    ('note', '※ 段付きフランジ（ジョグリング）は1つの製品に凹・凸・直線が同居する。'
             '一律のクリアランスでは形にならないため、区間ごとにクリアランスを変える。'),

    ('head', '■ 3. 片側クリアランス c ＝ 係数 × ta'),
    ('table', [
        ['区分', '軟鋼 SPCC', 'ステン SUS436L/429M', '高張力鋼', '狙い・注意'],
        ['凹・穴／普通バーリング', '1.00 ta', '1.00 ta', '1.00 ta',
         '標準。縁は成形中に約 0.7ta まで薄くなる'],
        ['凹・穴／しごきバーリング', '0.65〜0.75 ta', '0.65〜0.75 ta', '0.75〜0.85 ta',
         '高く均一なフランジ。一般文献は 0.60〜0.70t'],
        ['凹・穴／高フランジ狙い', '0.50〜0.65 ta', '0.60〜0.70 ta', '—',
         '肉薄・割れ注意。高張力は無理させない'],
        ['凹・開いた縁（伸びフランジ）', '1.00〜1.05 ta', '1.00〜1.05 ta', '1.05〜1.15 ta',
         '詰めない。薄くなる縁をしごくと割れる'],
        ['直線', '1.00〜1.05 ta', '1.00〜1.05 ta', '1.05〜1.15 ta', '標準。まずここから'],
        ['直線／しごき曲げ（角度優先）', '0.90〜0.98 ta', '0.92〜1.00 ta', '—',
         'SBは減るが加工力増・かじり。高張力は不可'],
        ['凸（縮みフランジ）', '1.05〜1.15 ta', '1.05〜1.15 ta', '1.10〜1.25 ta',
         '肉の逃げ場を作る。詰めるとしわを噛んで打痕'],
    ]),
    ('note', '※ ステン列は当社の主力材である HFS436L・HFS429M（フェライト系）を基準とする。'
             '当社の使用実績は 436L が最多、次いで 429M・409L で、SUS304 は少数。'),
    ('note', '※ フェライト系は SUS304 に比べ n値が小さく（0.25 対 0.34）引張強さも低い'
             '（482 対 652 MPa／JFEカタログ代表値）。加工硬化とスプリングバックが小さいため、'
             'クリアランスは 304 のように広げず軟鋼とほぼ同じでよい。'),
    ('note', '※ ただしフェライト系は伸びが小さい（31〜37% 対 304 の 56%）ため、高フランジの'
             '無理は 304 ほど利かない。穴拡げ率は 304 と同等（37〜38%）なので、普通〜しごき'
             'バーリング自体は問題なく成立する。'),
    ('note', '※ SUS304（オーステナイト系）を使う場合は、しごきバーリング 0.70〜0.80 ta、'
             '開いた縁・直線 1.05〜1.10 ta、凸 1.10〜1.20 ta と上表より一段広げる。'
             'しごき曲げは加工硬化・かじりのため不可。'),
    ('note', '※ 穴（閉じた円）は全周が拘束されるためしごきが成立し、フランジ高さと内径精度を'
             '稼げる。開いた縁は拘束がなく、しごくと割れを助長するため ta を下回らせない。'),
    ('note', '※ 凸でしわが出る場合の対策は二択。①型で潰す＝c を 1.00〜1.05 ta に詰めて軽く'
             'しごく（打痕・かじりのリスクを取る）②ブランクで逃がす＝c は広いままブランク形状で'
             '肉を減らす。どちらの方針かを設計時に決めておく。'),
    ('note', '※ 高張力鋼は延性が低く先端割れしやすい → 凹側はクリアランスを標準より '
             '5〜10% 大きめに。'),

    ('head', '■ 4. 付帯条件'),
    ('table', [
        ['項目', '標準', '備考'],
        ['曲げダイR（縁曲げ）', '3 × ta 程度',
         '小：焼付き・ショックライン・直角度不安定／大：材料が板厚方向へ逃げて曲がらない'],
        ['バーリングパンチ先端R', '0.5〜1.0 × ta',
         '平底・小Rほどフランジは高く出るが加工力は最大'],
        ['最小フランジ高さ', 'ダイR ＋ ta 以上', 'ダイRにブランク端がかかると曲がらない'],
        ['最小フランジ高さ（しごき曲げ）', '1.5 × ta', '拘束が効くため低い立ち上がりも可'],
        ['パッド押さえ力', 'ウェブ面積から決める',
         '弱いとウェブがパンチ方向へ引かれ寸法が動く（縁曲げの寸法ばらつきの主因）'],
        ['スプリングバック', 'クリアランス小・ダイR小で減る',
         'かじりと相反。高張力はオーバーベンド／リストライクで取る'],
    ]),

    ('head', '■ 5. 下穴径の逆算（凹・穴のみ）　※数値は計算例'
             '（d ≒ Dp − 2×(h − 0.43ta)、実型はトライで微調整）'),
    ('calc', [
        ['パンチ径 Dp (mm)', 10],
        ['基準板厚 ta (mm)', 1.6],
        ['必要フランジ高さ h (mm)', 3],
        ['推奨下穴径 d (mm)', None],          # None = 数式（下で組み立て）
    ]),
    ('note', '※ 下穴のカエリ（バリ）はフランジ立ち上がり外側（ダイ側）へ向ける。'
             '破断面を内側にすると割れ起点になりやすい。'),
    ('note', '※ タップ下穴用は、成形後のフランジ内径がネジ有効径を確保できるようパンチ径を'
             '選定する。しごきバーリングはフランジが薄くなるため、ねじ部の強度不足に注意。'),
    ('note', '※ 高フランジは1工程で無理せず、下穴→予備絞り→バーリングの複数工程分割も検討する。'),
]

NCOL = 5          # B〜F
COL0 = 2          # B列から書く


# 書式は他シート（ピアス・抜き／曲げ）と同じ定義を明示的に持つ。
# 座標から採取すると再実行時にレイアウトが変わって崩れるため、値で持つ。
JP = 'Noto Sans CJK SC'      # 見出し・和文
NUM = 'Arial'                # 数値欄
NAVY, BLUE = 'FF1F3864', 'FF2E5496'
GRAY, WHITE, INPUT_BG, INPUT_FG = 'FF555555', 'FFFFFFFF', 'FFFFF2CC', 'FF0000FF'


def _styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    thin = Side(style='thin', color='FFBFBFBF')
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    none = Border()
    plain = PatternFill()

    def st(font, fill=plain, border=none, align=None):
        return dict(font=font, fill=fill, border=border, alignment=align)

    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left', vertical='center', wrap_text=True)
    return {
        'title': st(Font(name=JP, size=12, bold=True, color=WHITE),
                    PatternFill('solid', fgColor=NAVY), align=ctr),
        'sub': st(Font(name=JP, size=9, color=GRAY), align=lft),
        'head': st(Font(name=JP, size=12, bold=True, color=WHITE),
                   PatternFill('solid', fgColor=BLUE), align=lft),
        'note': st(Font(name=JP, size=9, color=GRAY), align=lft),
        'th': st(Font(name=JP, size=10, bold=True, color=WHITE),
                 PatternFill('solid', fgColor=BLUE), box, ctr),
        'td_label': st(Font(name=JP, size=10, bold=True), plain, box, lft),
        'td_value': st(Font(name=NUM, size=10), plain, box, ctr),
        'td_note': st(Font(name=JP, size=10), plain, box, lft),
        'calc_label': st(Font(name=JP, size=10, bold=True), plain, box, lft),
        'calc_input': st(Font(name=NUM, size=10, bold=True, color=INPUT_FG),
                         PatternFill('solid', fgColor=INPUT_BG), box, ctr),
    }


def apply(cell, s):
    cell.font, cell.fill = copy(s['font']), copy(s['fill'])
    cell.border = copy(s['border'])
    if s['alignment'] is not None:
        cell.alignment = copy(s['alignment'])


def write_sheet(ws, S):
    r = 2
    calc_rows = {}
    for kind, payload in CONTENT:
        if kind in ('title', 'sub', 'head', 'note'):
            c = ws.cell(r, COL0, payload)
            apply(c, S[kind])
            ws.merge_cells(start_row=r, start_column=COL0,
                           end_row=r, end_column=COL0 + NCOL - 1)
            ws.row_dimensions[r].height = {'title': 27.75, 'sub': 25.5,
                                           'head': 19.4, 'note': 15.0}[kind]
            r += 2 if kind in ('title', 'head') else 1
            if kind == 'sub':
                r += 1
        elif kind == 'table':
            for i, row in enumerate(payload):
                for j, v in enumerate(row):
                    c = ws.cell(r, COL0 + j, v)
                    if i == 0:
                        apply(c, S['th'])
                    elif j == 0:
                        apply(c, S['td_label'])
                    elif j == len(row) - 1 and len(row) > 2:
                        apply(c, S['td_note'])
                    else:
                        apply(c, S['td_value'])
                ws.row_dimensions[r].height = 25.5 if i == 0 else 15.0
                r += 1
            r += 1
        elif kind == 'calc':
            for label, val in payload:
                calc_rows[label] = r
                apply(ws.cell(r, COL0, label), S['calc_label'])
                c = ws.cell(r, COL0 + 1)
                apply(c, S['calc_input'])
                c.value = val
                ws.row_dimensions[r].height = 15.0
                r += 1
            # d ＝ Dp − 2×(h − 0.43×ta)
            dp = calc_rows['パンチ径 Dp (mm)']
            ta = calc_rows['基準板厚 ta (mm)']
            h = calc_rows['必要フランジ高さ h (mm)']
            d = calc_rows['推奨下穴径 d (mm)']
            ws.cell(d, COL0 + 1).value = f'=C{dp}-2*(C{h}-0.43*C{ta})'
            r += 1
    return r


def bake_formula_cache(path):
    """Excelで開いて保存し直し、数式の計算結果をキャッシュへ焼き込む。

    openpyxl が書いた数式には計算結果が入らないため、PDF/PPTX ビルダー側
    (data_only=True で読む) では空セルに見えてしまう。Excel に一度計算させて
    値をキャッシュさせることで、数式を残したまま両方で正しく出る。
    """
    try:
        import win32com.client
    except ImportError:
        print('warn   : pywin32 が無いため数式の計算値を焼き込めません'
              '（Excelで開いて保存すれば解消します）')
        return
    excel = None
    try:
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        book = excel.Workbooks.Open(os.path.abspath(path))
        book.Save()
        book.Close(SaveChanges=False)
        print('baked  : 数式の計算値をキャッシュへ書き込み')
    except Exception as e:            # Excel が無い環境でも処理は続行する
        print(f'warn   : Excelでの再保存に失敗（{e}）。'
              f'Excelで開いて保存すれば解消します')
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def main():
    # Excelで開いたままだと保存が競合するため先に止める
    for path in (XLSX, KYOTSU):
        lock = os.path.join(os.path.dirname(path), '~$' + os.path.basename(path))
        if os.path.exists(lock):
            raise SystemExit(
                f'Excelで開いたままです: {os.path.basename(path)}\n'
                f'  （ロック: {lock}）\n'
                f'  Excelを閉じてから再実行してください。')

    stamp = date.today().strftime('%Y%m%d')
    for path in (XLSX, KYOTSU):
        bak = path.replace('.xlsx', f'.bak_フランジ追加前_{stamp}.xlsx')
        if not os.path.exists(bak):
            shutil.copy2(path, bak)
            print('backup :', os.path.basename(bak))

    # ---- クリアランス基準書 --------------------------------------------
    wb = openpyxl.load_workbook(XLSX)
    # 初回は「バーリング」、再実行時は「フランジ成形」を置き換える（冪等）
    src = OLD_SHEET if OLD_SHEET in wb.sheetnames else NEW_SHEET
    if src not in wb.sheetnames:
        raise SystemExit(f'元シートが見つかりません: {OLD_SHEET} / {NEW_SHEET}')
    old = wb[src]
    S = _styles()
    idx = wb.sheetnames.index(src)
    widths = {k: v.width for k, v in old.column_dimensions.items()}
    show_grid = old.sheet_view.showGridLines

    del wb[src]
    ws = wb.create_sheet(NEW_SHEET, idx)
    ws.sheet_view.showGridLines = show_grid
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    # 追加列（D・E）は数値欄なので中くらいの幅に揃える
    for col, w in (('D', 18.0), ('E', 18.0)):
        if col not in widths:
            ws.column_dimensions[col].width = w

    last = write_sheet(ws, S)
    wb.save(XLSX)
    print(f'updated: {os.path.basename(XLSX)} … 「{NEW_SHEET}」({last - 1}行)')
    bake_formula_cache(XLSX)

    # ---- 共通編（C5-2 / C5-3 / はじめに）--------------------------------
    wb2 = openpyxl.load_workbook(KYOTSU)
    ws2 = wb2['01_共通編']
    hdr = [c.value for c in ws2[1]]
    col_no = hdr.index('No') + 1
    col_body = hdr.index('規定内容') + 1
    add = {
        'C5-2': 'クリアランスは早見表「フランジ成形」による。基準板厚は素材板厚 t0 ではなく、'
                'その部位の実板厚 ta を使う（ドロー・張出しで引けている部位は t0 より薄い）。',
        'C5-3': 'クリアランスは早見表「フランジ成形」による。曲げ線が凹円弧（伸びフランジ）か'
                '凸円弧（縮みフランジ）かで必要クリアランスは逆になるため、区間ごとに判別して'
                '設定する。バーリングのクリアランス（板厚以下）をそのまま流用しない。',
    }
    for row in ws2.iter_rows(min_row=2):
        no = row[col_no - 1].value
        if no in add:
            cell = row[col_body - 1]
            if 'フランジ成形' not in (cell.value or ''):
                cell.value = (cell.value or '').rstrip() + ' ' + add[no]
                print(f'updated: 共通編 {no}')
    ws3 = wb2['00_はじめに']
    for row in ws3.iter_rows():
        for c in row:
            if c.value and 'ピアス・抜き／バーリング／曲げ' in str(c.value):
                c.value = str(c.value).replace('ピアス・抜き／バーリング／曲げ',
                                               'ピアス・抜き／フランジ成形／曲げ')
                print('updated: 共通編 00_はじめに')
    wb2.save(KYOTSU)


if __name__ == '__main__':
    main()
