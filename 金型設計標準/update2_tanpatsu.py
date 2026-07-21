#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""HMS単発編Excel：第2回更新

対応する要望:
  1. 7-6抜きダイの「（7-2参照）」が追いにくい → 参照先の内容を要約して併記
  2. 図面を他社資料から参考にしたい → 「参考図」列を追加し、該当ファイル名を記載
  3. BUR狙い値・CUT-BE設計値管理表は「数値」ではなく「計算方法・考え方」を織り込む
     → 第9章「設計値の決め方」を新設

使い方:
    python update2_tanpatsu.py 入力.xlsx 出力.xlsx

依存: pip install openpyxl
"""
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

WRAP = Alignment(wrap_text=True, vertical='top')
F9 = Font(name='Meiryo UI', size=9)
YELLOW = PatternFill('solid', fgColor='FFFFF2CC')
NEWROW = PatternFill('solid', fgColor='FFE8F0E8')
THIN = Side(style='thin', color='FFB0B8C4')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── 1. 相互参照を自己完結させる ────────────────────────
RULES = {
    '7-6': '抜きダイ（ダイプレート）は次による。'
           '・サイズは250mm×250mm以下とする。超える場合は刃先部を全て入れ子にするか、'
           '段差を付けて研磨面を小さくする（保守作業の軽減）　'
           '・取付ボルトの座ぐりは、ボルト頭部の深さよりさらに3mm以上深く座ぐる　'
           '・製品などの逃がしは研磨かさ上げを考慮し3mm以上深くする　'
           '・重量20kg以上の場合は吊下げ用タップ穴（M8以上）を2ヶ所以上設置　'
           '・角部は全てC1以上の面取り　'
           '・丸孔・異形孔は可能な限り規格品のボタンダイを使用し、軽圧入のはめあいとする。'
           'ツバ部の座ぐりはツバ厚＋3mm深くする　'
           '・入れ子には組み間違い防止を施す'
           '（コーナの面取を大きくする／取付ネジの位置をずらす／R・Lで形状を変える等。詳細は7-2）',
}

# ── 2. 参考図（他社資料のどのページを見るか） ──────────────
FIGS = {
    '2-1': '単発 P004（金型の表示方法）',
    '2-2': '単発 P005（ダイハイト・総重量）',
    '2-4': '単発 P006（材料の送り方向）',
    '2-5': '単発 P007（部品の材質、表面処理）',
    '2-6': '単発 P007（表面処理）',
    '3-1': '単発 P009（ガイドポスト）',
    '3-2': '単発 P008（ストリッパガイドピン＆ブシュ）',
    '3-3': '単発 P008',
    '3-4': '単発 P010（上型・U溝）',
    '3-6': '単発 P012〜P014（下型）',
    '4-1': '単発 P019（ワークのセット方向）',
    '4-2': '単発 P020（製品ガイド・クリアランス図）',
    '4-3': '単発 P021（ガイドの厚み・テーパ加工）',
    '4-4': '単発 P020〜P021（逆セット防止）',
    '4-5': '単発 P021（目視確認用の加工）',
    '4-6': '単発 P021（固定・材質）',
    '4-7': '単発 P020（パイロットピン 固定式・可動式）',
    '4-8': '単発 P020（先端が鋭利な部品の禁止）',
    '4-9': '単発 P018（アテピン 形状と高さ表）',
    '5-2': '単発 P025（スクラップカッター）',
    '5-3': '単発 P026（スクラップシュート）',
    '5-4': '単発 P027（カス詰まり対策）',
    '6-1': '単発 P015（金型位置決め）',
    '6-2': '単発 P016（ストロークエンドブロック）',
    '6-3': '単発 P029〜P030（吊りボルト・吊りフック）',
    '7-1': '単発 P035（入れ子にする部位）',
    '7-2': '単発 P035（組み間違い防止の4方法）',
    '7-3': '単発 P036（取外しタップ・座ぐり）',
    '7-4': '単発 P036（シムの可否）',
    '7-6': '単発 P037〜P038（抜きダイ）',
    '7-7': '単発 P039〜P040（抜きパンチ）',
    '7-8': '単発 P031（ノックピン）',
    '7-10': '単発 P033（コイルスプリング）',
    '7-12': '単発 P042（作業穴・逃がし穴）',
    '8-1': '単発 P043（下型安全カバー）',
    '8-2': '単発 P028（安全ボルト）',
}

# ── 3. 第9章 設計値の決め方（新設） ────────────────────
NEW_ITEMS = [
    ('9. 設計値の決め方', '9-1', '設計値管理表を作る',
     '金型ごとに設計値管理表を作成し、狙い値と実測値を記録する。'
     '要点は「図面値そのままで設計しない」こと。'
     '製品図の寸法に対し、スプリングバック・スプリングゴー・肉の逃げを見込んだ'
     '狙い値をパンチ側・ダイ側それぞれに設定する。'
     '記録する項目：品番／品名／材質／板厚／図面値／管理箇所（内径・外形など）／'
     '製品実測値／パンチ狙い値／ダイ狙い値／前工程／角R',
     '自社の★BUR,FL-DOWN狙い値・CUT-BE設計値管理表の構成をそのまま標準化',
     '', '', '品質'),

    ('9. 設計値の決め方', '9-2', 'バーリング（BUR）の狙い値',
     '製品の要求寸法に対し、パンチ・ダイの狙い値を別々に設定する。'
     '管理箇所（内径か外形か）を明記し、製品端面からの距離（パンチ側）と'
     'パンチ角Rを記録する。'
     '実測値は上下限の2点（例 128.0／128.4）で記録し、狙い値の妥当性を検証する。',
     '自社実績：HSS430CU t2.5 でΦ80±0.2 の内径管理など',
     '★バーリング花咲き・ネック（RBB コーン）', '', '品質'),

    ('9. 設計値の決め方', '9-3', 'フランジダウン（FL-DOWN）の狙い値',
     '径・巾それぞれについて、パンチ側とダイ側の狙い値を設定する。'
     'ダイ角Rと前工程（BLから一発フォームか、前工程ありか）を必ず記録する。'
     '前工程の有無で必要な狙い値が変わるため。',
     '自社実績：Φ128±0.2内径→パンチ-0.5／ダイ+0.3／ダイ角R5.0（BLから一発フォーム）'
     '、127.2切欠部→パンチ-0.4／ダイ-0.3',
     '', '⑪', '品質'),

    ('9. 設計値の決め方', '9-4', '抜きクリアランスの決め方',
     'クリアランスは板厚に対する％で決める。'
     'クリアランス（mm）＝ 板厚t × クリアランス率（％）。'
     '※率は材質・板厚・要求バリ高さで変える。'
     'せん断面の見た目より、<b>顧客要求のバリ高さを優先</b>すること。',
     '自社実績：t1.0／クリアランス8％＝0.08mm', '', '①②③④', '品質'),

    ('9. 設計値の決め方', '9-5', '材質係数と寿命の見込み',
     '材質ごとに係数を定め、金型の想定寿命（ショット数）とあわせて'
     '刃物の材質・表面処理を選定する。'
     '設計時に想定ショット数を明記し、メンテナンス周期の根拠とする。',
     '自社実績：SHGA270D-45 材質係数35／50万回想定',
     '★TOX耐用ショット数オーバー（6Y0）→交換頻度を再設定した実績あり', '', '保守'),

    ('9. 設計値の決め方', '9-6', 'ストローク・パスラインの設計値',
     'D.H.（ダイハイト）とP.L.（パスライン）を先に決め、'
     '上型・下型それぞれのST（ストローク量）を割り付ける。'
     '下型はUP/DOWNの可動量、上型は通常時と安全時の2値を持たせる。'
     'パンチとパッドの片側クリアランスも設計値として明記する。',
     '自社実績：D.H.350／P.L.220／下型ST25（UP25・DOWN0）／'
     '上型ST19（安全29）／パンチ-パッド片側0.4mm', '', '', '品質'),

    ('9. 設計値の決め方', '9-7', 'スプリングの選定',
     'リフター・パッドのスプリングは、線図から「安全重量」と「取付荷重」を求め、'
     '<b>取付荷重が安全重量を超えないこと</b>を確認する。'
     'パッドは周長と効率から必要荷重を算出する。'
     '選定根拠を設計値管理表に残すこと。',
     '自社のCUT-BE設計値管理表（リフター・パッド シート）の考え方',
     '', '⑬', '品質'),

    ('9. 設計値の決め方', '9-8', '実績のフィードバック',
     'トライ・量産で得た実測値を設計値管理表へ書き戻す。'
     '狙い値と実測値の差が繰り返し出る場合は、次の金型の狙い値に反映する。'
     '※この積み重ねが当社の設計基準そのものになる。',
     '', '', '', '品質'),
]


def build(src, dst):
    wb = openpyxl.load_workbook(src)
    ws = wb['01_単発編']

    # 参考図の列を追加（既存の最終列の後ろ）
    ncol = ws.max_column
    fig_col = ncol + 1
    ws.cell(1, fig_col, '参考図（他社資料）')
    ws.cell(1, fig_col).font = Font(name='Meiryo UI', size=9, bold=True, color='FFFFFFFF')
    ws.cell(1, fig_col).fill = PatternFill('solid', fgColor='FF2E5C9A')
    ws.cell(1, fig_col).alignment = Alignment(wrap_text=True, vertical='center',
                                              horizontal='center')
    ws.cell(1, fig_col).border = BOX
    ws.column_dimensions[get_column_letter(fig_col)].width = 30

    idx = {}
    for r in range(2, ws.max_row + 1):
        no = ws.cell(r, 2).value
        if no:
            idx[str(no).strip()] = r

    n_fig = 0
    for no, fig in FIGS.items():
        r = idx.get(no)
        if not r:
            continue
        c = ws.cell(r, fig_col, fig)
        c.alignment, c.font, c.border = WRAP, F9, BOX
        n_fig += 1

    n_rule = 0
    for no, txt in RULES.items():
        r = idx.get(no)
        if not r:
            continue
        c = ws.cell(r, 4, txt)
        c.alignment, c.font, c.fill = WRAP, F9, YELLOW
        ws.row_dimensions[r].height = max(34, min(150, len(txt) * 0.52))
        n_rule += 1

    # 第9章を末尾に追加
    r = ws.max_row + 1
    prev = None
    for chap, no, name, rule, ref, tora, zu, cat in NEW_ITEMS:
        rule_txt = rule.replace('<b>', '').replace('</b>', '')
        vals = [chap if chap != prev else '', no, name, rule_txt, ref, tora, cat, '', '']
        # 列構成: 章 No 項目 規定 参考値 過去トラ 分類 採否 備考 ／ 参考図
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, v)
            c.alignment, c.border = WRAP, BOX
            c.font = Font(name='Meiryo UI', size=9, bold=(i == 1 and chap != prev))
            c.fill = NEWROW
        ws.cell(r, fig_col, '—').border = BOX
        ws.cell(r, fig_col).fill = NEWROW
        if tora:
            ws.cell(r, 6).font = Font(name='Meiryo UI', size=9, bold=True, color='FF9C0006')
        ws.row_dimensions[r].height = max(36, min(140, len(rule_txt) * 0.52))
        prev = chap
        r += 1

    # 変更履歴へ追記
    if '03_変更履歴' in wb.sheetnames:
        ws2 = wb['03_変更履歴']
        rr = ws2.max_row + 1
        for d, k, v in [
            ('2026-07-22', '7-6 抜きダイ', '「（7-2参照）」だけでは追いにくいため、'
                                        '組み間違い防止の具体策を要約して併記'),
            ('2026-07-22', '参考図 列を追加', '他社資料のどのページに図があるかを記載。'
                                          '_図面参考フォルダに画像を書き出し済み（単発46枚・順送78点）'),
            ('2026-07-22', '第9章を新設', 'BUR狙い値・CUT-BE設計値管理表から'
                                       '「計算方法・考え方」を8項目として織り込み'),
        ]:
            for c, val in enumerate((d, k, v), 1):
                cell = ws2.cell(rr, c, val)
                cell.font, cell.alignment = F9, WRAP
            rr += 1

    wb.save(dst)
    print('作成しました:', dst)
    print('  参考図を記載: %d項目 ／ 規定を更新: %d項目 ／ 第9章を追加: %d項目'
          % (n_fig, n_rule, len(NEW_ITEMS)))


if __name__ == '__main__':
    build(sys.argv[1], sys.argv[2])
