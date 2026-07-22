#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""HMS 設計値管理表（記録様式）を作る

単発編 第9章／順送編 第12章「設計値の決め方」で使う記録表の実物。
技術部から「記録表は？」との要望を受けて作成。

シート構成
  00_使い方     … 何を、いつ、誰が書くか
  01_BUR        … バーリングの狙い値と実測値
  02_FL-DOWN    … フランジダウンの狙い値と実測値
  03_前工程基準  … ピアス面・トリム面を製品端面からいくつ下げる/上げるか
  04_抜き        … クリアランス・バリ高さ
  05_ストローク   … D.H./P.L./ST/クリアランス
  06_スプリング   … リフター・パッドの選定（たわみが許容内かを判定）
  07_金型台帳    … 1金型1行のサマリ

使い方:
    python make_sekkeichi_hyo.py 出力先.xlsx

依存: pip install openpyxl
"""
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = 'FF1F3864'
BLUE = 'FF2E5C9A'
TEAL = 'FF1C6B63'
LGRAY = 'FFF2F4F8'
AMBER = 'FFFDF6E3'
GREEN = 'FFEEF6EE'
PINK = 'FFFCEBEB'

THIN = Side(style='thin', color='FFB0B8C4')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
CTR = Alignment(wrap_text=True, vertical='center', horizontal='center')
F9 = Font(name='Meiryo UI', size=9)
F9B = Font(name='Meiryo UI', size=9, bold=True)
FH = Font(name='Meiryo UI', size=9, bold=True, color='FFFFFFFF')


def head(ws, cols, widths, fill=BLUE, row=1):
    for i, (c, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row, i, c)
        cell.fill = PatternFill('solid', fgColor=fill)
        cell.font = FH
        cell.alignment = CTR
        cell.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row + 1, 1)


def blank_rows(ws, ncol, start, count, fills=None):
    """記入用の空行に枠線を引く"""
    for r in range(start, start + count):
        for c in range(1, ncol + 1):
            cell = ws.cell(r, c)
            cell.border = BOX
            cell.font = F9
            cell.alignment = WRAP
            if fills and c in fills:
                cell.fill = PatternFill('solid', fgColor=fills[c])
        ws.row_dimensions[r].height = 22


def build(out):
    wb = Workbook()

    # ═══════════ 00 使い方 ═══════════
    ws = wb.active
    ws.title = '00_使い方'
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 104
    ws['A1'] = 'HMS 設計値管理表'
    ws['A1'].font = Font(name='Meiryo UI', size=16, bold=True, color=NAVY)
    ws['A2'] = '橋本工業株式会社　技術部'
    ws['A2'].font = Font(name='Meiryo UI', size=9, color='FF666666')

    notes = [
        ('この表の目的',
         '金型ごとに「狙い値」と「実測値」を記録し、次の金型の設計にいかすための表。'
         '図面値そのままで設計せず、スプリングバック・スプリングゴー・肉の逃げを見込んだ'
         '狙い値を必ず記録する。'),
        ('いつ書くか',
         '①設計時 … 狙い値を記入（この時点では実測値は空欄）　'
         '②トライ後 … 実測値を記入し、狙い値との差を確認　'
         '③量産開始後 … 安定した時点の実測値を記入して確定'),
        ('誰が書くか',
         '設計者が狙い値を、トライ担当が実測値を記入する。'
         '1つの金型につき1枚（1行）を作る。'),
        ('なぜ必要か',
         '狙い値と実測値の差が繰り返し出る場合、その傾向が当社の設計基準になる。'
         'この積み重ねが無いと、毎回ゼロから調整することになり、'
         '同じ手戻りを繰り返す。'),
        ('シートの使い分け',
         '01_BUR … バーリング　'
         '02_FL-DOWN … フランジダウン　'
         '03_前工程基準 … ピアス面・トリム面の下げ量／上げ量　'
         '04_抜き … クリアランスとバリ高さ　'
         '05_ストローク … D.H./P.L./ST　'
         '06_スプリング … リフター・パッドの選定　'
         '07_金型台帳 … 全金型のサマリ'),
        ('色の意味',
         '青系の見出し＝記入項目　'
         '黄色のセル＝設計時に記入　'
         '緑色のセル＝トライ・量産後に記入'),
    ]
    r = 4
    for t, b in notes:
        ws.cell(r, 1, t).font = Font(name='Meiryo UI', size=10, bold=True, color=NAVY)
        ws.cell(r, 1).alignment = WRAP
        c = ws.cell(r, 2, b)
        c.font, c.alignment = F9, WRAP
        ws.row_dimensions[r].height = 46
        r += 1

    # ═══════════ 01 BUR ═══════════
    ws = wb.create_sheet('01_BUR')
    cols = ['金型No', '品番', '品名', '材質', '板厚\nt', '図面値', '管理\n(内径/外形)',
            'パンチ\n狙い値', 'ダイ\n狙い値', 'パンチ角\nR', '前工程',
            '製品実測値\n(下限)', '製品実測値\n(上限)', '差異と気づき', '記入日']
    widths = [10, 22, 14, 12, 6, 12, 9, 9, 9, 8, 14, 11, 11, 30, 10]
    head(ws, cols, widths)
    blank_rows(ws, len(cols), 2, 30,
               fills={8: AMBER, 9: AMBER, 10: AMBER, 12: GREEN, 13: GREEN})
    ws.cell(33, 1, '※狙い値は設計時（黄）、実測値はトライ後（緑）に記入する。'
                   'パンチ側とダイ側は別々に設定すること。').font = F9

    # ═══════════ 02 FL-DOWN ═══════════
    ws = wb.create_sheet('02_FL-DOWN')
    cols = ['金型No', '品番', '品名', '材質', '板厚\nt', '図面値', '管理\n(径/巾)',
            'パンチ\n狙い値', 'ダイ\n狙い値', 'ダイ角\nR', '前工程\n(BLから一発/他)',
            '製品実測値\n(下限)', '製品実測値\n(上限)', '差異と気づき', '記入日']
    head(ws, cols, widths)
    blank_rows(ws, len(cols), 2, 30,
               fills={8: AMBER, 9: AMBER, 10: AMBER, 12: GREEN, 13: GREEN})
    ws.cell(33, 1, '※前工程の有無で必要な狙い値が変わるため、前工程欄を必ず記入する。').font = F9

    # ═══════════ 03 前工程基準（新規・要望対応） ═══════════
    ws = wb.create_sheet('03_前工程基準')
    ws.cell(1, 1, '前工程の面をどこに置くか　― 製品端面からの下げ量／上げ量')
    ws.cell(1, 1).font = Font(name='Meiryo UI', size=12, bold=True, color=NAVY)
    ws.cell(2, 1, 'バーリング前のピアス面、フランジダウン前のトリム面を、'
                  '製品端面からいくつずらすかの基準。ここがずれると成形後の寸法が出ない。')
    ws.cell(2, 1).font = F9
    ws.cell(2, 1).alignment = WRAP
    ws.merge_cells('A2:N2')
    ws.row_dimensions[2].height = 30

    cols = ['区分', '金型No', '品番', '材質', '板厚\nt', '成形後の\n要求寸法',
            '前工程の\n面の位置', '製品端面\nからの量', '基準の考え方',
            '成形後\n実測値', '差異', '判定', '備考', '記入日']
    widths = [12, 10, 20, 12, 6, 11, 11, 10, 26, 10, 8, 8, 22, 10]
    head(ws, cols, widths, row=4)
    blank_rows(ws, len(cols), 5, 26,
               fills={7: AMBER, 8: AMBER, 10: GREEN, 11: GREEN})
    for i, v in enumerate(['BUR前ピアス', 'BUR前ピアス', 'FL-DOWN前トリム', 'FL-DOWN前トリム'], 5):
        ws.cell(i, 1, v).font = F9
    ws.cell(33, 1, '※区分は「BUR前ピアス（下げ量）」「FL-DOWN前トリム（上げ量）」を記入。'
                   '成形で材料が動く分を見込んで前工程の面を決める。'
                   '実績が溜まったら板厚・材質ごとの標準値を定める。').font = F9
    ws.merge_cells('A33:N33')
    ws.row_dimensions[33].height = 30

    # ═══════════ 04 抜き ═══════════
    ws = wb.create_sheet('04_抜き')
    cols = ['金型No', '品番', '材質', '板厚\nt', 'クリアランス率\n(%)',
            'クリアランス\n(mm)', '要求バリ高さ', 'せん断長さ\n狙い',
            '実測バリ高さ', '判定', '想定ショット数', '材質係数', '備考', '記入日']
    widths = [10, 22, 12, 6, 12, 11, 11, 10, 11, 8, 12, 9, 26, 10]
    head(ws, cols, widths)
    blank_rows(ws, len(cols), 2, 26,
               fills={5: AMBER, 6: AMBER, 8: AMBER, 9: GREEN, 10: GREEN})
    ws.cell(29, 1, '※クリアランス(mm) ＝ 板厚t × クリアランス率(%)。'
                   'せん断面の見た目より、顧客要求のバリ高さを優先する。').font = F9

    # ═══════════ 05 ストローク ═══════════
    ws = wb.create_sheet('05_ストローク')
    cols = ['金型No', '品番', 'プレス', 'D.H.\n(mm)', 'P.L.\n(mm)',
            '下型ST\n(mm)', '下型UP', '下型DOWN', '上型ST\n(mm)', '上型ST\n(安全)',
            'パンチ-パッド\n片側クリア', '材料幅', 'ピッチ', '備考', '記入日']
    widths = [10, 22, 10, 8, 8, 8, 7, 8, 8, 9, 12, 8, 8, 24, 10]
    head(ws, cols, widths)
    blank_rows(ws, len(cols), 2, 26, fills={4: AMBER, 5: AMBER, 6: AMBER, 9: AMBER, 11: AMBER})
    ws.cell(29, 1, '※D.H.とP.L.を先に決め、上型・下型のSTを割り付ける。'
                   '上型は通常時と安全時の2値を持たせる。').font = F9

    # ═══════════ 06 スプリング ═══════════
    ws = wb.create_sheet('06_スプリング')
    ws.cell(1, 1, 'スプリングの選定　― 「たわみ量が許容内か」を確認する')
    ws.cell(1, 1).font = Font(name='Meiryo UI', size=12, bold=True, color=NAVY)
    ws.cell(2, 1,
            '考え方：スプリングは「縮められる限界（許容たわみ）」が決まっている。'
            '取付時に既に縮めている量（取付たわみ）＋加工中に縮む量（作動たわみ）が、'
            'その限界を超えないことを確認する。'
            '超えると密着して折れる、または荷重が急上昇して金型を傷める。')
    ws.cell(2, 1).font = F9
    ws.cell(2, 1).alignment = WRAP
    ws.merge_cells('A2:M2')
    ws.row_dimensions[2].height = 44

    cols = ['金型No', '用途\n(リフター/パッド)', '品番・型番', '本数',
            '自由長\n(mm)', '許容たわみ\n(mm)', '取付たわみ\n(mm)', '作動たわみ\n(mm)',
            'たわみ合計\n(mm)', '判定\n(合計≦許容)', '必要荷重\n(kgf)',
            '発生荷重\n(kgf)', '備考']
    widths = [10, 14, 18, 6, 9, 11, 11, 11, 11, 11, 10, 10, 26]
    head(ws, cols, widths, row=4)
    blank_rows(ws, len(cols), 5, 24,
               fills={6: LGRAY, 7: AMBER, 8: AMBER, 9: GREEN, 10: GREEN})
    # 判定の数式を入れておく
    for r in range(5, 29):
        ws.cell(r, 9, '=IF(AND(G%d<>"",H%d<>""),G%d+H%d,"")' % (r, r, r, r)).font = F9
        ws.cell(r, 10, '=IF(OR(I%d="",F%d=""),"",IF(I%d<=F%d,"OK","超過"))' % (r, r, r, r)).font = F9B
    ws.cell(30, 1, '※許容たわみはメーカーカタログの値を転記する。'
                   'パッドの必要荷重は「製品の周長×効率」から求める。'
                   '判定が「超過」なら、より長いスプリングにするか本数を増やす。').font = F9
    ws.merge_cells('A30:M30')
    ws.row_dimensions[30].height = 30

    # ═══════════ 07 金型台帳 ═══════════
    ws = wb.create_sheet('07_金型台帳')
    cols = ['金型No', '品番', '品名', '客先', '型種\n(単発/順送)', '設計者',
            '製作年月', 'プレス', '想定ショット数', '実績ショット数',
            '最終メンテ日', '次回メンテ予定', '設計値管理表\n記入済', '備考']
    widths = [10, 22, 16, 14, 10, 10, 10, 10, 12, 12, 11, 12, 12, 26]
    head(ws, cols, widths, fill=TEAL)
    blank_rows(ws, len(cols), 2, 40, fills={13: PINK})
    ws.cell(44, 1, '※金型1つにつき1行。「設計値管理表 記入済」欄に'
                   '各シートへの記入状況（BUR/FL/抜き 等）を書く。').font = F9

    wb.save(out)
    print('作成しました:', out)
    print('  シート:', ', '.join(wb.sheetnames))


if __name__ == '__main__':
    build(sys.argv[1] if len(sys.argv) > 1 else 'HMS設計値管理表.xlsx')
