#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""順送・単発を整合し、共通する項目を共通編へ移す（第2次移行）

第1次移行(migrate_to_kyotsu.py)のあと、順送・単発を突き合わせて残っていた
「型種によらず同じ」項目を共通編へ集約する。共通編の文言は両編を1つに整合。

■ やること
  1. 共通編に新項目を追加（両編を整合した本文）
  2. 単発・順送から移した項目を削除
  3. 順送5-7は「組み間違い防止」の一文だけ抜いて残す
  4. 単発9-9は共通編C5-4と同一なので削除（重複解消）
  5. 各章末の「→ 共通編を参照」行に、新しく移したC項目を追記

依存: pip install openpyxl
"""
import os
import re
import warnings

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

warnings.filterwarnings('ignore')

BASE = os.path.join(os.path.expanduser('~'), 'Desktop', '橋本_作業データ', '金型仕様書')
WRAP = Alignment(wrap_text=True, vertical='top')
F9 = Font(name='Meiryo UI', size=9)
THIN = Side(style='thin', color='FFB0B8C4')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
REFFILL = PatternFill('solid', fgColor='FFE1EFDA')

# ---- 共通編へ追加する項目（章, No, 項目, 規定内容, 移行元）
NEW_COMMON = [
    ('1. 総則', 'C1-6', '提出物',
     '型初品時は、評価用サンプル・初品検査成績表・測定データを品質管理へ提示し、評価を受ける。'
     '提出物・数量は客先指示がある場合はそれに従う。'
     '※売型をせず全て社内生産のため、金型納入の手続きは行わない。',
     '単発1-5,1-6／順送1-6'),
    ('2. 金型の基本', 'C2-5', '金型総重量',
     '金型の総重量は、使用するホイストの吊り上げ荷重を超えないこと。上限は2.8トン未満とする。',
     '順送2-5（単発2-3）'),
    ('2. 金型の基本', 'C2-6', '上型の取付（U溝）',
     '上型ホルダにはU溝（取付け溝）を設ける。当社はオートクランプが無いため、U溝にφ80ワッシャと'
     'ボルトで固定する。U溝の長さは30mm以上、クランプ部の厚みは30〜50mm。U溝の位置と数は、'
     '110トン以下＝2ヶ所（右奥・左手前）、150トン以上＝4ヶ所。U溝の幅・ピッチは金型取付標準表に'
     '合わせ、ピッチ公差は±1.0mm以内。φ80ワッシャが他部品と干渉しないこと。'
     '上型ホルダの厚みはプレス能力に応じて確保する。',
     '単発3-4／順送4-1'),
    ('2. 金型の基本', 'C2-7', '下型',
     '下型の最下部に取付位置決めを設ける。順送・ロボットライン等では取付位置決め穴（Φ22±0.2）を'
     'あけ、U溝を設ける。大型プレス・ロボットラインでは取付位置決めプレートを配置する。',
     '単発3-6／順送4-2'),
    ('2. 金型の基本', 'C2-8', '工程の分割',
     '抜き（分断を含む）・曲げ・絞りなど、役割の異なる工程は分割する。',
     '単発3-7／順送4-5'),
    ('3. 部品の共通仕様', 'C3-9', '可動パッド',
     '可動パッドは、ストリッパプレートに固定されたパンチが偏心・摩耗するのを防ぐために設ける。'
     '板厚が厚い（1.2mm以上が目安）と抜き荷重が大きく、サブガイドピンの片減りが起きやすいため。'
     '該当する金型があるか確認のうえ採否を判断する。',
     '順送4-6'),
    ('3. 部品の共通仕様', 'C3-10', '組み間違い防止',
     '左右対称形状の入れ子や刻印パンチには組み間違い防止を施す。'
     '①1つのコーナの面取を大きくし、挿入側も同様にする　②取付ネジの位置をずらす　'
     '③R/Lテーキンは面取位置・大きさ・形状を変える　④ツバ付パンチは回り止めの面数を変える。',
     '単発7-2／順送5-7'),
    ('4. 安全', 'C4-5', '前面ガード・安全カバー',
     '段取り時の手挟み等を防ぐため、金型前面に前面ガード（安全カバー）を設置する。'
     '下型プレートが可動する構造で10mm以上の隙間がある場合は必ず取付ける'
     '（押し上げバネを内蔵する場合は不要）。カバー用プレートはクッションの可動が確認できること。',
     '単発8-1／順送10-2'),
    ('6. 排出（共通）', 'C6-4', '製品の払出し',
     '加工後の製品が金型上に残らない構造とする。自重で滑り落ちる角度・形状とし（勾配R10以上）、'
     '引っ掛かりをなくす。原則は送り方向へ排出し、困難な場合は承認を得て作業者側（前面排出）と'
     'してよい。スクラップは型外へ確実に排出するシュートを設ける。'
     '※滑り出し角度は製品形状・重量・油量で変わるため、うまく払い出せている金型を実測して'
     '標準値を定める（未定）。',
     '単発5-1,5-3／順送7-5,7-7'),
]

# ---- 各編で削除する項目 No → その項目が属する章の見出し語（参照行の判定用）と、対応するC項目
# (No, C項目表示)  ※Noだけで章は行から判定する
TAN_MOVE = {  # 単発編
    '1-5': 'C1-6 提出物', '1-6': 'C1-6 提出物',
    '2-3': 'C2-5 金型総重量',
    '3-4': 'C2-6 上型の取付（U溝）', '3-6': 'C2-7 下型', '3-7': 'C2-8 工程の分割',
    '5-1': 'C6-4 製品の払出し', '5-3': 'C6-4 製品の払出し',
    '7-2': 'C3-10 組み間違い防止',
    '8-1': 'C4-5 前面ガード・安全カバー',
    '9-9': 'C5-4 前工程の面の位置',   # 重複解消（C5-4と同一）
}
JUN_MOVE = {  # 順送編
    '1-6': 'C1-6 提出物',
    '2-5': 'C2-5 金型総重量',
    '4-1': 'C2-6 上型の取付（U溝）', '4-2': 'C2-7 下型',
    '4-5': 'C2-8 工程の分割', '4-6': 'C3-9 可動パッド',
    '7-5': 'C6-4 製品の払出し', '7-7': 'C6-4 製品の払出し',
    '10-2': 'C4-5 前面ガード・安全カバー',
}
# 順送5-7は削除せず、組み間違いの一文だけ抜く。参照にはC3-10を足す
JUN_EDIT = {
    '5-7': ('入れ子はボルト等で固定されていること（圧入禁止）（量産時の振動対策）。'
            '抜きダイの切刃部入れ子には取り外し作業用タップ穴を設置する。',
            'C3-10 組み間違い防止'),
}


def rebuild_common():
    f = os.path.join(BASE, 'HMS金型設計標準_共通編.xlsx')
    wb = openpyxl.load_workbook(f)
    ws = wb['01_共通編']
    ncol = ws.max_column
    rows = []
    cur = ''
    for r in range(2, ws.max_row + 1):
        c1 = str(ws.cell(r, 1).value or '').strip()
        if c1:
            cur = c1
        no = str(ws.cell(r, 2).value or '').strip()
        if not no:
            continue
        vals = [ws.cell(r, c).value for c in range(1, ncol + 1)]
        rows.append({'chap': cur, 'no': no, 'vals': vals})
    # 章順
    order = []
    for x in rows:
        if x['chap'] not in order:
            order.append(x['chap'])
    # 新項目を章ごとにまとめる
    add = {}
    for chap, no, item, rule, src in NEW_COMMON:
        add.setdefault(chap, []).append((no, item, rule, src))
        if chap not in order:
            order.append(chap)
    # 書き直し
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
    r = 2
    prev = None
    for ch in order:
        items = [x for x in rows if x['chap'] == ch]
        for x in items:
            for i, v in enumerate(x['vals'], 1):
                cell = ws.cell(r, i, v if i != 1 else (ch if ch != prev else ''))
                cell.alignment, cell.border = WRAP, BOX
                cell.font = Font(name='Meiryo UI', size=9, bold=(i == 1 and ch != prev))
            prev = ch
            r += 1
        for (no, item, rule, src) in add.get(ch, []):
            ws.cell(r, 1, ch if ch != prev else '')
            ws.cell(r, 2, no)
            ws.cell(r, 3, item)
            ws.cell(r, 4, rule)
            ws.cell(r, 10, src)         # 移行元
            for c in range(1, ncol + 1):
                cell = ws.cell(r, c)
                cell.alignment, cell.border = WRAP, BOX
                cell.font = Font(name='Meiryo UI', size=9, bold=(c == 1 and ch != prev))
            ws.row_dimensions[r].height = max(30, min(140, len(rule) * 0.5))
            prev = ch
            r += 1
    wb.save(f)
    print('共通編: %d項目を追加' % len(NEW_COMMON))


def parse_refs(text):
    """参照行テキストから C番号 の一覧を、表示（Cx-y 項目名）ごとに返す"""
    body = re.sub(r'^.*による：\s*', '', str(text or ''))
    parts = [p.strip() for p in re.split(r'[、，]', body) if p.strip()]
    return parts


def rebuild_edition(fname, sheet, move, edit):
    f = os.path.join(BASE, fname)
    wb = openpyxl.load_workbook(f)
    ws = wb[sheet]
    ncol = ws.max_column
    rows = []
    cur = ''
    for r in range(2, ws.max_row + 1):
        c1 = str(ws.cell(r, 1).value or '').strip()
        if c1:
            cur = c1
        no = str(ws.cell(r, 2).value or '').strip()
        if not no:
            continue
        vals = [ws.cell(r, c).value for c in range(1, ncol + 1)]
        rows.append({'chap': cur, 'no': no, 'vals': vals})
    order = []
    for x in rows:
        if x['chap'] not in order:
            order.append(x['chap'])

    # 章ごとの既存参照 + 新規追加参照
    ref_by_chap = {}
    for x in rows:
        if x['no'] == '→':
            ref_by_chap[x['chap']] = parse_refs(x['vals'][3])
    # 移動・編集で足す参照
    for x in rows:
        if x['no'] in move:
            ref_by_chap.setdefault(x['chap'], [])
            disp = move[x['no']]
            if disp not in ref_by_chap[x['chap']]:
                ref_by_chap[x['chap']].append(disp)
        if x['no'] in edit:
            ref_by_chap.setdefault(x['chap'], [])
            disp = edit[x['no']][1]
            if disp not in ref_by_chap[x['chap']]:
                ref_by_chap[x['chap']].append(disp)

    # 残す項目（移動されず、参照行でもない）。編集対象は本文差し替え
    dec_col = 8 if sheet == '01_単発編' else 9
    keep = []
    for x in rows:
        if x['no'] == '→':
            continue
        if x['no'] in move:
            continue
        if x['no'] in edit:
            x['vals'][3] = edit[x['no']][0]   # 規定内容を差し替え
        keep.append(x)

    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
    r = 2
    prev = None
    n_del = 0
    for ch in order:
        items = [x for x in keep if x['chap'] == ch]
        refs = ref_by_chap.get(ch, [])
        if not items and not refs:
            continue
        for x in items:
            for i, v in enumerate(x['vals'], 1):
                cell = ws.cell(r, i, v if i != 1 else (ch if ch != prev else ''))
                cell.alignment, cell.border = WRAP, BOX
                cell.font = Font(name='Meiryo UI', size=9, bold=(i == 1 and ch != prev))
            ws.row_dimensions[r].height = max(28, min(150, len(str(x['vals'][3] or '')) * 0.5))
            prev = ch
            r += 1
        if refs:
            txt = '次の項目は【共通編】による： ' + '、'.join(refs)
            ws.cell(r, 1, ch if ch != prev else '')
            ws.cell(r, 2, '→')
            ws.cell(r, 3, '共通編を参照')
            ws.cell(r, 4, txt)
            for c in range(1, ncol + 1):
                cell = ws.cell(r, c)
                cell.alignment, cell.border, cell.fill = WRAP, BOX, REFFILL
                cell.font = Font(name='Meiryo UI', size=9, bold=(c == 1 and ch != prev))
            ws.row_dimensions[r].height = max(24, min(90, len(txt) * 0.5))
            prev = ch
            r += 1
    n_del = len(move) + len(edit)  # 参考
    wb.save(f)
    keep_n = len(keep)
    print('%s: 残%d項目、削除%d、編集%d' % (sheet, keep_n, len(move), len(edit)))


if __name__ == '__main__':
    rebuild_common()
    rebuild_edition('HMS金型設計標準_単発編_骨格.xlsx', '01_単発編', TAN_MOVE, {})
    rebuild_edition('HMS金型設計標準_順送編_骨格.xlsx', '01_順送編', JUN_MOVE, JUN_EDIT)
