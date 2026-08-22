# -*- coding: utf-8 -*-
"""金型設計クリアランス基準書：ステンレスの基準材を SUS304 →
フェライト系（SUS436L / SUS429M）へ差し替える。

背景
  当社のステンレス使用実績は HFS436L が最多（次いで HFS429M・HFS409L）で、
  SUS304 は少数。基準書の「ステン」列が SUS304 だと現場の主力材と合わない。

  フェライト系とオーステナイト系(304)の差（JFEスチール 自動車用ステンレス鋼
  カタログ 代表値, 板厚0.8mm）:
    SUS436L : 引張482MPa  伸び31%  n値0.25  r値1.4  穴拡げ率37.5%
    SUS304  : 引張652MPa  伸び57%  n値0.34  r値1.0  穴拡げ率38.4%
  → 加工硬化・スプリングバックは 304 より小さく、軟鋼に近い。
    一方で伸びが小さく、急な曲げ・高フランジの無理は利かない。

  「フランジ成形」シートは add_flange_clearance.py 側で対応済み。
  本スクリプトは残る「ピアス・抜き」「曲げ」の2シートを対象とする。

  ⚠ 曲げの最小曲げ内R・スプリングバック角は上記の材料特性からの当社暫定値。
    トライ実績が溜まったら C5-9 のとおり書き戻すこと。

使い方: python update_stainless_grade.py
"""
import os

import openpyxl

import build_hms_docs as B

XLSX = B.CLEAR_XLSX

# シート名 → [(検索する先頭セル文字列, [置換後の行の値...]), ...]
ROW_EDITS = {
    'ピアス・抜き': [
        ('ステンレス (SUS304)',
         ['ステン (SUS436L/429M)', '5〜7%', '7〜9%', '9〜11%', '7.5%',
          'フェライト系。304より加工硬化小（n値0.25・引張482MPa）']),
    ],
    '曲げ': [
        ('最小曲げ内R(対板厚)', None, 3, '1.0〜2.5t'),   # 3列目(ステン列)だけ差し替え
        ('スプリングバック目安', None, 3, '2〜4°'),
    ],
}

# 見出しセルの置換（列名）
HEADER_EDITS = {
    'ピアス・抜き': [('ステンレス (SUS304)', 'ステン (SUS436L/429M)')],
    '曲げ': [('ステン SUS304', 'ステン SUS436L/429M')],
}

# 追記する注記（既にあれば追記しない）
ADD_NOTES = {
    'ピアス・抜き': [
        '※ ステン列は当社主力の HFS436L・HFS429M（フェライト系）を基準とする。'
        'SUS304（オーステナイト系）は加工硬化が大きく（n値0.34・引張652MPa）、'
        '6〜8% / 7〜10% / 10〜12%（中央値8%）と一段大きめに取る。',
    ],
    '曲げ': [
        '※ ステン列は当社主力の HFS436L・HFS429M（フェライト系）を基準とする。'
        'フェライト系は延性が低く（伸び31% 対 SUS304の57%）急な曲げで割れやすいため'
        '曲げ内Rを大きめに、加工硬化が小さい（n値0.25 対 0.34）ためスプリングバックは'
        '小さい。SUS304 の場合は 最小曲げ内R 1.0〜2.0t・スプリングバック 3〜6°。',
        '※ フェライト系は圧延方向に筋状の凹凸（リジング）が出ることがある。'
        '外観面は圧延方向を考慮して板取りする。',
        '※ ステン列の曲げ内R・スプリングバック角は材料特性からの暫定値。'
        'トライ実績が溜まったら実測値へ更新する（C5-9）。',
    ],
}


def note_style(ws):
    """同じシート内の既存の ※ 行から書式を借りる。"""
    from copy import copy
    for r in range(ws.max_row, 1, -1):
        c = ws.cell(r, 2)
        if isinstance(c.value, str) and c.value.startswith('※'):
            return r, dict(font=copy(c.font), alignment=copy(c.alignment))
    return None, None


def main():
    lock = os.path.join(os.path.dirname(XLSX), '~$' + os.path.basename(XLSX))
    if os.path.exists(lock):
        raise SystemExit(f'Excelで開いたままです: {os.path.basename(XLSX)}\n'
                         f'  Excelを閉じてから再実行してください。')

    wb = openpyxl.load_workbook(XLSX)
    for sheet, edits in ROW_EDITS.items():
        ws = wb[sheet]
        for edit in edits:
            key = edit[0]
            target = None
            for row in ws.iter_rows(min_col=2, max_col=2):
                if str(row[0].value).strip() == key:
                    target = row[0].row
                    break
            if target is None:
                print(f'skip   : {sheet} 「{key}」が見つかりません')
                continue
            if len(edit) == 2:                      # 行まるごと差し替え
                for j, v in enumerate(edit[1]):
                    ws.cell(target, 2 + j).value = v
            else:                                   # 特定列だけ差し替え
                _, _, col, val = edit
                ws.cell(target, 1 + col).value = val
            print(f'updated: {sheet} 行「{key}」')

    for sheet, pairs in HEADER_EDITS.items():
        ws = wb[sheet]
        for old, new in pairs:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.strip() == old:
                        c.value = new
                        print(f'updated: {sheet} 見出し「{old}」→「{new}」')

    for sheet, notes in ADD_NOTES.items():
        ws = wb[sheet]
        existing = {str(c.value) for row in ws.iter_rows() for c in row
                    if isinstance(c.value, str)}
        last, style = note_style(ws)
        if last is None:
            print(f'skip   : {sheet} 注記の書式が見つかりません')
            continue
        r = last
        width = 6 if sheet == 'ピアス・抜き' else 5    # B..G / B..F
        for text in notes:
            if any(text[:24] in e for e in existing):
                continue
            r += 1
            c = ws.cell(r, 2, text)
            c.font, c.alignment = style['font'], style['alignment']
            ws.merge_cells(start_row=r, start_column=2,
                           end_row=r, end_column=1 + width)
            ws.row_dimensions[r].height = 15.0
            print(f'added  : {sheet} 注記 …{text[:26]}')

    wb.save(XLSX)
    print('saved  :', os.path.basename(XLSX))


if __name__ == '__main__':
    main()
