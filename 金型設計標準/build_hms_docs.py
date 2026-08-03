#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""HMS金型設計標準の3編を、清書したPDF／PowerPointにする

Excelはドラフト用（作業メモ・採否・過去トラ等の欄を持つ）。ここではそこから
「規定内容」だけを取り出して、社内で読む・外注へ渡せる体裁に清書する。

■ 出さないもの（Excelの作業欄）
  備考（担当者の疑問メモ）／過去トラ／相関図／移行元／要検討 は最終文書に載せない。
  採否＝対象外・不要 の項目は本文から外し、巻末に「適用対象外」として一覧で残す
  （黙って消すと、後で「なぜ無いのか」が分からなくなるため）。

■ クリアランス基準
  共通編の第5章（設計値の決め方）に本文として織り込む。
  別ブック「金型設計クリアランス基準書.xlsx」の数表をそのまま章内へ差し込む。

使い方:
    python build_hms_docs.py            # 3編すべて pdf/ へ出力
    python build_hms_docs.py 共通         # 共通編だけ

依存: pip install openpyxl reportlab python-pptx
"""
import os
import re
import sys
import warnings

import openpyxl

warnings.filterwarnings('ignore')

BASE = r'C:\Users\Owner\Desktop\金型仕様書'
OUT = os.path.join(BASE, '出力')
FIGDIR = os.path.join(BASE, 'HMS図')
CLEAR_XLSX = os.path.join(BASE, '金型設計クリアランス基準書.xlsx')

# (Excelファイル, シート名, 表示名, 図ファイルの接頭辞)
EDITIONS = {
    '共通': ('HMS金型設計標準_共通編.xlsx', '01_共通編', '共通編', '共'),
    '単発': ('HMS金型設計標準_単発編_骨格.xlsx', '01_単発編', '単発編', '単'),
    '順送': ('HMS金型設計標準_順送編_骨格.xlsx', '01_順送編', '順送編', '順'),
}

# クリアランス基準書のうち本文へ入れるシート（表紙は除く）
CLEAR_SHEETS = ['ピアス・抜き', 'バーリング', '曲げ', '順送レイアウト', '設計チェックリスト']


# ================================================================ 読み込み
def _cols(hdr):
    def find(*names):
        for n in names:
            if n in hdr:
                return hdr.index(n) + 1
        return None
    return {
        'no': find('No'),
        'item': find('項目'),
        'rule': find('規定内容', '規定内容（案）'),
        'dec': find('採否'),
    }


def load_edition(path, sheet):
    """[(章, [item...]), ...] を返す。item は dict。"""
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = _cols(hdr)
    chapters = []
    cur_name = None
    cur_items = None
    chap = ''
    for r in range(2, ws.max_row + 1):
        c1 = str(ws.cell(r, 1).value or '').strip()
        if c1:
            chap = c1
        no = str(ws.cell(r, col['no']).value or '').strip()
        if not no:
            continue
        if chap != cur_name:
            cur_name = chap
            cur_items = []
            chapters.append((chap, cur_items))
        if no == '→':                     # 共通編を参照する行
            ref = str(ws.cell(r, 4).value or '')
            cur_items.append({'kind': 'ref', 'text': ref})
            continue
        cur_items.append({
            'kind': 'item',
            'no': no,
            'item': str(ws.cell(r, col['item']).value or ''),
            'rule': str(ws.cell(r, col['rule']).value or ''),
            'dec': str(ws.cell(r, col['dec']).value or '') if col['dec'] else '',
        })
    return chapters


def load_intro(path):
    ws = openpyxl.load_workbook(path, data_only=True)['00_はじめに']
    rows = []
    for r in range(2, ws.max_row + 1):
        a = str(ws.cell(r, 1).value or '').strip()
        b = str(ws.cell(r, 2).value or '').strip()
        if a or b:
            rows.append((a, b))
    return rows


def load_clearance():
    """クリアランス基準書の各シートを {シート名: block列} で返す。
    block = ('title'|'sub'|'head'|'note'|'para', text) または ('table', [row...])"""
    wb = openpyxl.load_workbook(CLEAR_XLSX, data_only=True)
    out = {}
    for sh in CLEAR_SHEETS:
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        rows = []
        for r in range(1, ws.max_row + 1):
            vals = [('' if ws.cell(r, c).value in (None, '') else
                     str(ws.cell(r, c).value).strip())
                    for c in range(1, ws.max_column + 1)]
            rows.append(vals)
        out[sh] = _parse_clear_sheet(rows)
    return out


def _parse_clear_sheet(rows):
    blocks = []
    tbl = []
    seen_title = False

    def flush():
        nonlocal tbl
        if tbl:
            w = max((len(r) for r in tbl), default=0)
            grid = [r + [''] * (w - len(r)) for r in tbl]
            # 右端の全空列を落とす
            while w > 1 and all(row[w - 1] == '' for row in grid):
                w -= 1
                grid = [row[:w] for row in grid]
            # 左端の全空列を落とす（元Excelの空白A列対策）
            while w > 1 and all(row[0] == '' for row in grid):
                grid = [row[1:] for row in grid]
                w -= 1
            blocks.append(('table', grid))
            tbl = []

    for vals in rows:
        filled = [i for i, v in enumerate(vals) if v != '']
        if not filled:
            flush()
            continue
        if len(filled) == 1:
            flush()
            t = vals[filled[0]]
            if not seen_title:
                blocks.append(('title', t))
                seen_title = True
            elif t.startswith('■'):
                blocks.append(('head', t.lstrip('■ ').strip()))
            elif t.startswith('※'):
                blocks.append(('note', t.lstrip('※ ').strip()))
            else:
                blocks.append(('sub', t))
        else:
            tbl.append([vals[i] for i in range(filled[-1] + 1)])
    flush()
    return blocks


# ================================================================ PDF
def build_pdf(chapters, intro, title_ja, clear_section, checklist, prefix, out_path):
    # clear_section = (barタイトル, 説明文, [(シート名, blocks), ...]) または None
    #                 …「設計値」章の直後に差し込む早見表群
    # checklist     = (見出し, blocks) または None …巻末（付録）へ入れる
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import (BaseDocTemplate, Frame, Image, KeepTogether,
                                    NextPageTemplate, PageBreak, PageTemplate,
                                    Paragraph, Spacer, Table, TableStyle)
    from reportlab.lib.utils import ImageReader

    pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
    F = 'HeiseiKakuGo-W5'
    NAVY = colors.HexColor('#1F3864')
    TEAL = colors.HexColor('#1C6B63')
    GRAY = colors.HexColor('#555555')
    LIGHT = colors.HexColor('#EAF0F6')
    GREENBG = colors.HexColor('#E1EFDA')
    LINE = colors.HexColor('#B7C2D0')

    def ps(name, **kw):
        kw.setdefault('fontName', F)
        return ParagraphStyle(name, **kw)

    def esc(s):
        # reportlab の Paragraph は <, &, > をタグとして解釈するため無害化する。
        # そのうえで改行だけ <br/> に戻す。
        s = str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        return s.replace('\n', '<br/>')

    st_cover_t = ps('ct', fontSize=24, leading=32, textColor=NAVY)
    st_cover_s = ps('cs', fontSize=13, leading=20, textColor=GRAY)
    st_cover_org = ps('co', fontSize=12, leading=18, textColor=colors.black)
    st_intro_h = ps('ih', fontSize=11, leading=16, textColor=NAVY)
    st_intro_b = ps('ib', fontSize=9.5, leading=15, textColor=colors.black)
    st_chap = ps('ch', fontSize=14, leading=18, textColor=colors.white)
    st_item = ps('it', fontSize=10.5, leading=14, textColor=NAVY)
    st_body = ps('bd', fontSize=9.5, leading=15, textColor=colors.black,
                 alignment=TA_LEFT)
    st_ref = ps('rf', fontSize=9, leading=13, textColor=TEAL)
    st_head = ps('hd', fontSize=10.5, leading=15, textColor=TEAL)
    st_sub = ps('sb', fontSize=9.5, leading=14, textColor=GRAY)
    st_note = ps('nt', fontSize=8.5, leading=12, textColor=GRAY)
    st_tc = ps('tc', fontSize=8.5, leading=11.5, textColor=colors.black)
    st_th = ps('th', fontSize=8.5, leading=11.5, textColor=colors.white)
    st_appx = ps('ax', fontSize=9, leading=13, textColor=colors.black)
    st_foot = ps('ft', fontSize=8, leading=10, textColor=GRAY)
    from reportlab.lib.enums import TA_CENTER
    st_cap = ps('cap', fontSize=9, leading=12, textColor=GRAY, alignment=TA_CENTER)

    story = []

    # ---- 表紙
    story.append(Spacer(1, 55 * mm))
    story.append(Paragraph('HMDS 金型設計標準', st_cover_t))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph('【%s】' % title_ja, st_cover_t))
    story.append(Spacer(1, 10 * mm))
    story.append(Paragraph('橋本工業株式会社　技術チーム', st_cover_org))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph('HASHIMOTO KOGYO — Mold Design Standard（HMDS）', st_cover_s))
    story.append(Spacer(1, 16 * mm))
    intro_tbl = []
    for a, b in intro:
        if a and b:
            intro_tbl.append([Paragraph(esc(a), st_intro_h), Paragraph(esc(b), st_intro_b)])
        elif a:
            intro_tbl.append([Paragraph(esc(a), st_intro_h), ''])
    if intro_tbl:
        t = Table(intro_tbl, colWidths=[34 * mm, 128 * mm])
        t.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LINEBELOW', (0, 0), (-1, -2), 0.4, LINE),
        ]))
        story.append(t)
    story.append(NextPageTemplate('body'))
    story.append(PageBreak())

    # ---- 本文
    def chapter_bar(name):
        t = Table([[Paragraph(name, st_chap)]], colWidths=[162 * mm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), NAVY),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        return t

    def render_table(rows, first_is_header=True):
        data = []
        for i, r in enumerate(rows):
            style = st_th if (first_is_header and i == 0) else st_tc
            data.append([Paragraph(esc(c), style) for c in r])
        ncol = max(len(r) for r in rows)
        # 1列目やや広め、残り均等
        avail = 162 * mm
        if ncol == 1:
            widths = [avail]
        else:
            first = avail * 0.26
            widths = [first] + [(avail - first) / (ncol - 1)] * (ncol - 1)
        t = Table(data, colWidths=widths, repeatRows=1 if first_is_header else 0)
        ts = [
            ('GRID', (0, 0), (-1, -1), 0.4, LINE),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F8FB')]),
        ]
        if first_is_header:
            ts.append(('BACKGROUND', (0, 0), (-1, 0), TEAL))
        t.setStyle(TableStyle(ts))
        return t

    def render_blocks(blocks):
        flow = []
        for kind, payload in blocks:
            if kind == 'title':
                flow.append(Spacer(1, 2 * mm))
                flow.append(Paragraph('◆ ' + esc(payload), st_item))
                flow.append(Spacer(1, 1 * mm))
            elif kind == 'sub':
                flow.append(Paragraph(esc(payload), st_sub))
            elif kind == 'head':
                flow.append(Spacer(1, 1 * mm))
                flow.append(Paragraph('■ ' + esc(payload), st_head))
            elif kind == 'note':
                flow.append(Paragraph('※ ' + esc(payload), st_note))
            elif kind == 'para':
                flow.append(Paragraph(esc(payload), st_body))
            elif kind == 'table':
                flow.append(Spacer(1, 1 * mm))
                flow.append(render_table(payload))
                flow.append(Spacer(1, 1 * mm))
        return flow

    def clearance_flowables():
        bar_title, intro_text, sheets = clear_section
        flow = [Spacer(1, 3 * mm), chapter_bar(bar_title), Spacer(1, 2 * mm),
                Paragraph(intro_text, st_body), Spacer(1, 2 * mm)]
        for sh, blocks in sheets:
            flow += render_blocks(blocks)
        return flow

    def figure_for(no):
        """項目番号に対応する図があれば、幅に収めた Image を返す。"""
        path = os.path.join(FIGDIR, 'hms_%s%s.png' % (prefix, no))
        if not os.path.exists(path):
            return None
        iw, ih = ImageReader(path).getSize()
        maxw = 168 * mm          # 本文幅いっぱいまで使う
        maxh = 180 * mm
        scale = min(maxw / iw, maxh / ih)
        return Image(path, width=iw * scale, height=ih * scale)

    appendix = []       # 適用対象外の一覧
    # この文字列を含む章は新しいページの先頭から始める（読みやすさのため）
    page_break_before = ('送り・ガイド',)
    for ci, (cname, items) in enumerate(chapters):
        block = []
        if ci > 0 and any(k in cname for k in page_break_before):
            block.append(PageBreak())
        block += [Spacer(1, 3 * mm), chapter_bar(cname), Spacer(1, 2 * mm)]
        for it in items:
            if it['kind'] == 'ref':
                block.append(Paragraph('▶ ' + esc(it['text']), st_ref))
                block.append(Spacer(1, 1 * mm))
                continue
            if it['dec'] in ('対象外', '不要'):
                appendix.append((it['no'], it['item'], it['dec']))
                continue
            head = esc('%s　%s' % (it['no'], it['item']))
            # 他社資料(高木)の整理番号Sxxxは標準に残さない。図は自社へ取り込み済み。
            rule = it['rule']
            rule = re.sub(r'※?\s*(図|表|写真)?は?他社資料\s*S\d+[〜～\-]*S?\d*[^。]*?(参照|あり)。?',
                          '', rule)
            rule = re.sub(r'（\s*S\d+[〜～\-]*S?\d*[^）]*）', '', rule)
            rule = re.sub(r'※?\s*S\d+\s*に写真あり。?', '', rule)
            body = esc(rule.strip())
            group = [Paragraph(head, st_item), Paragraph(body, st_body)]
            fig = figure_for(it['no'])
            if fig is not None:
                # 図番号キャプション（図1・図2…）は本文が参照しないため付けない。
                group += [Spacer(1, 1.5 * mm), fig]
            group.append(Spacer(1, 2.5 * mm))
            block.append(KeepTogether(group))
        story.extend(block)
        # 「設計値」章の直後に早見表群を差し込む（共通=クリアランス・曲げ／順送=順送レイアウト）
        if clear_section and '設計値' in cname:
            story.extend(clearance_flowables())

    # 巻末付録（「付. 適用対象外とした項目」「付. 設計チェックリスト」）は
    # 本標準では載せない方針のため出力しない。
    #   ・適用対象外：対象外/不要の項目は本文から外すだけとし、一覧化しない
    #   ・設計チェックリスト：早見表（本文中）に集約したため巻末には付けない
    # ※ appendix・checklist の受け渡しは互換のため残すが、描画はしない。
    _ = (appendix, checklist)

    # ---- レイアウト
    def on_page(canv, doc):
        canv.saveState()
        canv.setFont(F, 8)
        canv.setFillColor(GRAY)
        canv.drawString(20 * mm, 12 * mm,
                        'HMDS 金型設計標準【%s】　橋本工業株式会社' % title_ja)
        canv.drawRightString(190 * mm, 12 * mm, '%d' % doc.page)
        canv.setStrokeColor(LINE)
        canv.setLineWidth(0.4)
        canv.line(20 * mm, 15 * mm, 190 * mm, 15 * mm)
        canv.restoreState()

    def on_cover(canv, doc):
        canv.saveState()
        canv.setFillColor(NAVY)
        canv.rect(0, 272 * mm, 210 * mm, 25 * mm, fill=1, stroke=0)
        canv.restoreState()

    doc = BaseDocTemplate(out_path, pagesize=A4,
                          leftMargin=20 * mm, rightMargin=20 * mm,
                          topMargin=20 * mm, bottomMargin=18 * mm)
    fw = A4[0] - 40 * mm
    fh = A4[1] - 38 * mm
    frame = Frame(20 * mm, 18 * mm, fw, fh, id='f')
    doc.addPageTemplates([
        PageTemplate(id='cover', frames=[frame], onPage=on_cover),
        PageTemplate(id='body', frames=[frame], onPage=on_page),
    ])
    story.insert(0, NextPageTemplate('cover'))
    doc.build(story)
    return out_path


# ================================================================ 実行
def main():
    os.makedirs(OUT, exist_ok=True)
    keys = [sys.argv[1]] if len(sys.argv) > 1 else list(EDITIONS)
    cl = load_clearance()

    def sheets(names):
        return [(n, cl[n]) for n in names if n in cl]

    # 早見表の振り分け：共通=クリアランス・曲げ／順送=順送レイアウト／チェックリストは共通の巻末
    common_clear = ('5-補. クリアランス・曲げ 早見表',
                    'クリアランス（ピアス・抜き／バーリング）と曲げの早見表。'
                    '数値は標準値で、材料ロット・型構造・製品要求により調整する。',
                    sheets(['ピアス・抜き', 'バーリング', '曲げ']))
    junso_clear = ('補. 順送レイアウト早見表',
                   '順送型のレイアウト設計基準（早見表）。数値は標準値で、'
                   '製品・材料により調整する。',
                   sheets(['順送レイアウト']))
    checklist = ('付. 設計チェックリスト', cl['設計チェックリスト']) \
        if '設計チェックリスト' in cl else None

    CLEAR = {'共通': (common_clear, checklist),
             '単発': (None, None),
             '順送': (junso_clear, None)}
    for k in keys:
        fname, sheet, ja, prefix = EDITIONS[k]
        path = os.path.join(BASE, fname)
        chapters = load_edition(path, sheet)
        intro = load_intro(path)
        clear_section, chk = CLEAR[k]
        out = os.path.join(OUT, 'HMDS金型設計標準_%s.pdf' % ja)
        build_pdf(chapters, intro, ja, clear_section, chk, prefix, out)
        print('PDF:', out)


if __name__ == '__main__':
    main()
