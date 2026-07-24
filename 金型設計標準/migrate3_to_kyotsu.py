#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""第3次移行：ユーザーがM列にマークした部品系項目を共通編へ集約

単発7章・順送4-5章の部品/刃物系をユーザーがM列「共通」でマーク。
それを共通編へ移し、両編で重複する内容は1つに整合する。
また違う行の同内容（単発6-1↔C2-7、順送5-4↔C5-5）は既存項目へ統合する。

依存: pip install openpyxl
"""
import os
import re
import warnings

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

warnings.filterwarnings('ignore')
BASE = r'C:\Users\Owner\Desktop\金型仕様書'
WRAP = Alignment(wrap_text=True, vertical='top')
THIN = Side(style='thin', color='FFB0B8C4')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
REFFILL = PatternFill('solid', fgColor='FFE1EFDA')


def F(b=False):
    return Font(name='Meiryo UI', size=9, bold=b)


def rule_of(ws, no):
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    ci = (hdr.index('規定内容') + 1) if '規定内容' in hdr else (hdr.index('規定内容（案）') + 1)
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 2).value or '') == no:
            return str(ws.cell(r, ci).value or '')
    return ''


def load_edition_texts():
    twb = openpyxl.load_workbook(os.path.join(BASE, 'HMS金型設計標準_単発編_骨格.xlsx'), data_only=True)
    jwb = openpyxl.load_workbook(os.path.join(BASE, 'HMS金型設計標準_順送編_骨格.xlsx'), data_only=True)
    t = twb['01_単発編']
    j = jwb['01_順送編']
    return {'t': lambda no: rule_of(t, no), 'j': lambda no: rule_of(j, no)}


def build_new_common(src):
    """（章, No, 項目, 規定内容, 移行元）のリストを返す"""
    T, J = src['t'], src['j']
    tan77 = T('7-7')                      # 単発 抜きパンチ（詳細）
    tan74 = T('7-4')
    tan76 = T('7-6').replace('詳細は7-2', '詳細はC3-10')
    jun48 = J('4-8').replace('※プレート構成は 2-1 の図を参照', '※プレート構成はC2-9の図を参照')
    merged_punch = tan77 + '　・パンチホルダーを取り外すことなく脱着できること。'
    merged_insert_fix = (
        '入れ子はボルト等で固定する（圧入禁止。量産時の振動対策）。'
        '取り外し作業用タップ穴を設置する（締付ボルト穴にタップ加工でも可。'
        '抜きダイの切刃部入れ子にも設置）。'
        '組付ボルトの座ぐり深さは再研磨を考慮し頭部厚み＋3mm以上深くする。')
    return [
        ('2. 金型の基本', 'C2-10', 'パンチホルダー', J('4-7'), '順送4-7'),
        ('2. 金型の基本', 'C2-11', 'バッキングプレート', jun48, '順送4-8'),
        ('3. 部品の共通仕様', 'C3-11', '抜きダイ', tan76, '単発7-6'),
        ('3. 部品の共通仕様', 'C3-12', '抜きパンチ', merged_punch, '単発7-7／順送5-5'),
        ('3. 部品の共通仕様', 'C3-13', '入れ子にする部位', T('7-1'), '単発7-1'),
        ('3. 部品の共通仕様', 'C3-14', '入れ子の固定・取り外し', merged_insert_fix, '単発7-3／順送5-7'),
        ('3. 部品の共通仕様', 'C3-15', '座ぐりインサート部品', J('5-10'), '順送5-10'),
        ('3. 部品の共通仕様', 'C3-16', '標準パンチ・ダイの使用基準', J('5-6'), '順送5-6'),
        ('3. 部品の共通仕様', 'C3-17', '打刻テーキンパンチ', J('5-9'), '順送5-9'),
        ('3. 部品の共通仕様', 'C3-18', 'ディスタンススペーサ', J('5-11'), '順送5-11'),
        ('3. 部品の共通仕様', 'C3-19', 'シムによる調整', tan74, '単発7-4'),
        ('3. 部品の共通仕様', 'C3-20', '肉盛りによる調整', '肉盛り調整した部品は図面に反映する。', '単発7-5'),
        ('3. 部品の共通仕様', 'C3-21', 'ダイの刃先形状', J('5-1'), '順送5-1'),
        ('3. 部品の共通仕様', 'C3-22', 'ダイの逃がし', J('5-2'), '順送5-2'),
        ('3. 部品の共通仕様', 'C3-23', '切刃の面粗さ', J('5-3'), '順送5-3'),
    ]


# 既存の共通編項目へ追記（重複統合）
EDIT_COMMON = {
    'C2-7': '　大型プレスはΦ55のロケートピンを使用する。',
    'C5-5': '（例）SPC・板厚2.0mm → 2.0÷3≒0.7、せん断長さ0.7mm以上。',
}

# 削除して共通参照へ（No → 表示）
TAN_MOVE = {
    '6-1': 'C2-7 下型', '7-1': 'C3-13 入れ子にする部位', '7-3': 'C3-14 入れ子の固定・取り外し',
    '7-4': 'C3-19 シムによる調整', '7-5': 'C3-20 肉盛りによる調整',
    '7-6': 'C3-11 抜きダイ', '7-7': 'C3-12 抜きパンチ',
}
JUN_MOVE = {
    '4-7': 'C2-10 パンチホルダー', '4-8': 'C2-11 バッキングプレート',
    '5-1': 'C3-21 ダイの刃先形状', '5-2': 'C3-22 ダイの逃がし', '5-3': 'C3-23 切刃の面粗さ',
    '5-4': 'C5-5 抜きクリアランスの決め方', '5-5': 'C3-12 抜きパンチ',
    '5-6': 'C3-16 標準パンチ・ダイの使用基準', '5-7': 'C3-14 入れ子の固定・取り外し',
    '5-9': 'C3-17 打刻テーキンパンチ', '5-10': 'C3-15 座ぐりインサート部品',
    '5-11': 'C3-18 ディスタンススペーサ',
}


def rebuild_common(new_items):
    f = os.path.join(BASE, 'HMS金型設計標準_共通編.xlsx')
    wb = openpyxl.load_workbook(f)
    ws = wb['01_共通編']
    ncol = ws.max_column
    hdr = [ws.cell(1, c).value for c in range(1, ncol + 1)]
    rcol = (hdr.index('規定内容') + 1)
    rows, cur = [], ''
    for r in range(2, ws.max_row + 1):
        c1 = str(ws.cell(r, 1).value or '').strip()
        if c1:
            cur = c1
        no = str(ws.cell(r, 2).value or '').strip()
        if not no:
            continue
        vals = [ws.cell(r, c).value for c in range(1, ncol + 1)]
        # 既存項目への追記
        if no in EDIT_COMMON and EDIT_COMMON[no] not in str(vals[rcol - 1] or ''):
            vals[rcol - 1] = str(vals[rcol - 1] or '') + EDIT_COMMON[no]
        rows.append({'chap': cur, 'vals': vals})
    order = []
    for x in rows:
        if x['chap'] not in order:
            order.append(x['chap'])
    add = {}
    for chap, no, item, ruletxt, srcv in new_items:
        add.setdefault(chap, []).append((no, item, ruletxt, srcv))
    ws.delete_rows(2, ws.max_row - 1)
    r, prev = 2, None
    for ch in order:
        for x in [y for y in rows if y['chap'] == ch]:
            for i, v in enumerate(x['vals'], 1):
                cell = ws.cell(r, i, v if i != 1 else (ch if ch != prev else ''))
                cell.alignment, cell.border = WRAP, BOX
                cell.font = F(i == 1 and ch != prev)
            ws.row_dimensions[r].height = max(28, min(150, len(str(x['vals'][rcol - 1] or '')) * 0.5))
            prev = ch
            r += 1
        for (no, item, ruletxt, srcv) in add.get(ch, []):
            ws.cell(r, 1, ch if ch != prev else '')
            ws.cell(r, 2, no)
            ws.cell(r, 3, item)
            ws.cell(r, rcol, ruletxt)
            ws.cell(r, 10, srcv)
            for c in range(1, ncol + 1):
                cell = ws.cell(r, c)
                cell.alignment, cell.border = WRAP, BOX
                cell.font = F(c == 1 and ch != prev)
            ws.row_dimensions[r].height = max(30, min(150, len(ruletxt) * 0.5))
            prev = ch
            r += 1
    wb.save(f)
    print('共通編: +%d項目、既存%d件へ追記' % (len(new_items), len(EDIT_COMMON)))


def parse_refs(text):
    body = re.sub(r'^.*による：\s*', '', str(text or ''))
    return [p.strip() for p in re.split(r'[、，]', body) if p.strip()]


def rebuild_edition(fname, sheet, move):
    f = os.path.join(BASE, fname)
    wb = openpyxl.load_workbook(f)
    ws = wb[sheet]
    ncol = ws.max_column
    hdr = [ws.cell(1, c).value for c in range(1, ncol + 1)]
    rcol = (hdr.index('規定内容') + 1) if '規定内容' in hdr else (hdr.index('規定内容（案）') + 1)
    rows, cur = [], ''
    for r in range(2, ws.max_row + 1):
        c1 = str(ws.cell(r, 1).value or '').strip()
        if c1:
            cur = c1
        no = str(ws.cell(r, 2).value or '').strip()
        if not no:
            continue
        rows.append({'chap': cur, 'no': no,
                     'vals': [ws.cell(r, c).value for c in range(1, ncol + 1)]})
    order = []
    for x in rows:
        if x['chap'] not in order:
            order.append(x['chap'])
    ref_by_chap = {}
    for x in rows:
        if x['no'] == '→':
            ref_by_chap[x['chap']] = parse_refs(x['vals'][3])
    for x in rows:
        if x['no'] in move:
            ref_by_chap.setdefault(x['chap'], [])
            if move[x['no']] not in ref_by_chap[x['chap']]:
                ref_by_chap[x['chap']].append(move[x['no']])
    keep = [x for x in rows if x['no'] != '→' and x['no'] not in move]
    ws.delete_rows(2, ws.max_row - 1)
    r, prev = 2, None
    for ch in order:
        items = [x for x in keep if x['chap'] == ch]
        refs = ref_by_chap.get(ch, [])
        if not items and not refs:
            continue
        for x in items:
            for i, v in enumerate(x['vals'], 1):
                cell = ws.cell(r, i, v if i != 1 else (ch if ch != prev else ''))
                cell.alignment, cell.border = WRAP, BOX
                cell.font = F(i == 1 and ch != prev)
            ws.row_dimensions[r].height = max(28, min(150, len(str(x['vals'][rcol - 1] or '')) * 0.5))
            prev = ch
            r += 1
        if refs:
            txt = '次の項目は【共通編】による： ' + '、'.join(refs)
            ws.cell(r, 1, ch if ch != prev else '')
            ws.cell(r, 2, '→')
            ws.cell(r, 3, '共通編を参照')
            ws.cell(r, 4, txt)
            for c in range(1, ncol + 1):
                cell = ws.cell(r, c)
                cell.alignment, cell.border, cell.fill = WRAP, BOX, REFFILL
                cell.font = F(c == 1 and ch != prev)
            ws.row_dimensions[r].height = max(24, min(110, len(txt) * 0.5))
            prev = ch
            r += 1
    wb.save(f)
    print('%s: 残%d項目、移動%d' % (sheet, len(keep), len(move)))


if __name__ == '__main__':
    src = load_edition_texts()
    rebuild_common(build_new_common(src))
    rebuild_edition('HMS金型設計標準_単発編_骨格.xlsx', '01_単発編', TAN_MOVE)
    rebuild_edition('HMS金型設計標準_順送編_骨格.xlsx', '01_順送編', JUN_MOVE)
