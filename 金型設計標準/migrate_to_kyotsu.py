#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""共通編へ移した項目を、単発編・順送編から削除する

各編の該当行を削除し、代わりに章の先頭へ「共通編を参照」の1行を置く。
これにより重複がなくなり、1か所直せば両方に反映されるようになる。

■ 削除の判断
  共通編の「移行元」列に書かれた項目番号を各編から探して削除する。
  型種で内容が異なるため残す項目（上型・下型・送り方向・抜きパンチ・提出物）は
  共通編に移していないので、この処理では削除されない。

■ 安全策
  ・技術部が記入した備考・採否は、削除前に「99_削除した項目」シートへ退避する
  ・元ファイルは変更せず、別名で出力する

使い方:
    python migrate_to_kyotsu.py 共通編.xlsx 単発編.xlsx 順送編.xlsx 出力先フォルダ

依存: pip install openpyxl
"""
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

WRAP = Alignment(wrap_text=True, vertical='top')
CTR = Alignment(wrap_text=True, vertical='center', horizontal='center')
F9 = Font(name='Meiryo UI', size=9)
THIN = Side(style='thin', color='FFB0B8C4')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
REFFILL = PatternFill('solid', fgColor='FFE1EFDA')   # 共通編参照の行
TEAL = 'FF1C6B63'


def read_migration(kyotsu):
    """共通編の移行元列から {単発の番号: 共通編No, ...} を作る"""
    ws = openpyxl.load_workbook(kyotsu, data_only=True)['01_共通編']
    tan, jun, info = {}, {}, {}
    for r in range(2, ws.max_row + 1):
        cno = ws.cell(r, 2).value
        src = str(ws.cell(r, 10).value or '')
        name = ws.cell(r, 3).value
        if not cno or not src or src.startswith('（'):
            continue
        info[str(cno)] = name
        for part in src.split('／'):
            part = part.strip()
            if part.startswith('単発'):
                for x in part.replace('単発', '').split(','):
                    if x.strip():
                        tan[x.strip()] = str(cno)
            elif part.startswith('順送'):
                for x in part.replace('順送', '').split(','):
                    if x.strip():
                        jun[x.strip()] = str(cno)
    return tan, jun, info


def migrate(path, sheet, mapping, info, out):
    """行を消して作り直すのではなく、全行を読み出して並べ直す。
    行削除だと章見出し（章名は各章の先頭行にしか入っていない）が
    一緒に消えたり、参照行の挿入位置がずれたりするため。"""
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    ncol = ws.max_column
    dec_col = 8 if sheet == '01_単発編' else 9
    note_col = 9 if sheet == '01_単発編' else 10

    # 全行を (章, 各列の値) として読み出す。章名は前の行から引き継ぐ
    rows = []
    cur_chap = ''
    for r in range(2, ws.max_row + 1):
        v = str(ws.cell(r, 1).value or '').strip()
        if v:
            cur_chap = v
        vals = [ws.cell(r, c).value for c in range(1, ncol + 1)]
        no = str(ws.cell(r, 2).value or '').strip()
        if not no:
            continue
        rows.append({'chap': cur_chap, 'no': no, 'vals': vals,
                     'name': str(ws.cell(r, 3).value or ''),
                     'dec': str(ws.cell(r, dec_col).value or ''),
                     'note': str(ws.cell(r, note_col).value or '')})

    # 章の並び順を保持
    chap_order = []
    for x in rows:
        if x['chap'] not in chap_order:
            chap_order.append(x['chap'])

    targets = [x for x in rows if x['no'] in mapping]
    keep = [x for x in rows if x['no'] not in mapping]

    # 章ごとに、参照する共通編の項目をまとめる
    ref_by_chapter = {}
    for x in targets:
        ref_by_chapter.setdefault(x['chap'], []).append((mapping[x['no']], x['name']))

    # 中身を全部消して書き直す
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    r = 2
    added = 0
    prev_chap = None
    for ch in chap_order:
        items = [x for x in keep if x['chap'] == ch]
        refs = ref_by_chapter.get(ch, [])
        if not items and not refs:
            continue
        for x in items:
            for i, v in enumerate(x['vals'], 1):
                cell = ws.cell(r, i, v if i != 1 else (ch if ch != prev_chap else ''))
                cell.alignment, cell.border = WRAP, BOX
                cell.font = Font(name='Meiryo UI', size=9,
                                 bold=(i == 1 and ch != prev_chap))
            rule = str(x['vals'][3] or '')
            ws.row_dimensions[r].height = max(30, min(150, len(rule) * 0.5))
            prev_chap = ch
            r += 1
        if refs:
            txt = '次の項目は【共通編】による： ' + '、'.join(
                '%s %s' % (c, n) for c, n in refs)
            ws.cell(r, 1, ch if ch != prev_chap else '')
            ws.cell(r, 2, '→')
            ws.cell(r, 3, '共通編を参照')
            ws.cell(r, 4, txt)
            for c in range(1, ncol + 1):
                cell = ws.cell(r, c)
                cell.alignment, cell.border, cell.fill = WRAP, BOX, REFFILL
                cell.font = Font(name='Meiryo UI', size=9,
                                 bold=(c == 1 and ch != prev_chap))
            ws.row_dimensions[r].height = max(24, min(90, len(txt) * 0.5))
            prev_chap = ch
            added += 1
            r += 1

    targets = [(0, x['no'], x['chap'], x['name'], mapping[x['no']], x['dec'], x['note'])
               for x in targets]

    # 退避シート
    if '99_削除した項目' in wb.sheetnames:
        del wb['99_削除した項目']
    ws2 = wb.create_sheet('99_削除した項目')
    ws2.cell(1, 1, '共通編へ移したため、この編から削除した項目')
    ws2.cell(1, 1).font = Font(name='Meiryo UI', size=11, bold=True, color='FF1F3864')
    ws2.cell(2, 1, '記入されていた採否・備考は失わないようここに残しています。'
                   '内容は共通編へ反映済みか確認してください。')
    ws2.cell(2, 1).font = F9
    cols = ['旧No', '章', '項目', '共通編での番号', '記入されていた採否', '記入されていた備考']
    for i, (h, w) in enumerate(zip(cols, (8, 16, 26, 14, 14, 60)), 1):
        cell = ws2.cell(4, i, h)
        cell.fill = PatternFill('solid', fgColor=TEAL)
        cell.font = Font(name='Meiryo UI', size=9, bold=True, color='FFFFFFFF')
        cell.alignment, cell.border = CTR, BOX
        ws2.column_dimensions[get_column_letter(i)].width = w
    for r2, (_r, no, ch, name, cno, dec, note) in enumerate(targets, 5):
        for i, v in enumerate((no, ch, name, cno, dec, note), 1):
            cell = ws2.cell(r2, i, v)
            cell.font, cell.alignment, cell.border = F9, WRAP, BOX
        if note or dec:
            ws2.cell(r2, 6).fill = PatternFill('solid', fgColor='FFFFF2CC')

    wb.save(out)
    return len(targets), added


if __name__ == '__main__':
    if len(sys.argv) < 5:
        print(__doc__)
        sys.exit(1)
    kyotsu, tan_f, jun_f, outdir = sys.argv[1:5]
    os.makedirs(outdir, exist_ok=True)
    tan_map, jun_map, info = read_migration(kyotsu)
    print('共通編に集約された項目: 単発%d件 / 順送%d件' % (len(tan_map), len(jun_map)))
    print()
    n1, a1 = migrate(tan_f, '01_単発編', tan_map, info,
                     os.path.join(outdir, os.path.basename(tan_f)))
    print('単発編: %d項目を削除、%d章に参照行を追加' % (n1, a1))
    n2, a2 = migrate(jun_f, '01_順送編', jun_map, info,
                     os.path.join(outdir, os.path.basename(jun_f)))
    print('順送編: %d項目を削除、%d章に参照行を追加' % (n2, a2))
