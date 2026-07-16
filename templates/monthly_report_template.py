#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
月次会レポート ツール（Excel入力 → PDF生成）
================================================
【重要】このスクリプトにコードは置くが、実際の月次データ(xlsx/PDF)は
このリポジトリには置かない（GitHub Pagesは公開リポジトリのため）。
Excel入力ファイルはローカルPC（Desktop等）で管理する。

使い方:
  python monthly_report_template.py init  <xlsxパス>          … 空の雛形からExcel入力シートを新規作成
  python monthly_report_template.py build <xlsxパス> <PDFパス> … Excel入力シートを読み込みPDFを生成

【毎月の運用フロー】
  1. 前月分の完成済みxlsxをコピーして今月分のファイル名にする
     （部署構成やシート形式はほぼ毎月同じなので、initより「前月コピー」が基本）
  2. #月次会(Slack C06PJTWQJ15)の当月投稿と、深津英介氏添付のプレス実績/全体実績PDF、
     八木氏添付の苦情品質PDF、Z:ドライブの目的目標実施計画.xlsx を突き合わせて数値を確認
  3. Excelの各シートを更新（重要事項／品質実績～重要アクション）
  4. build コマンドでPDFを生成し、内容を確認（pypdf 等でテキスト抽出しての機械チェックも可）
  5. 大きな流れの詳細は docs/SYSTEMS.md の「月次会レポート自動生成」項を参照

初回・新規部署追加時のみ init で雛形を作成する。
シート構成は下部の SHEET SCHEMA を参照。
"""
import sys, os, io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, Image, HRFlowable, PageBreak)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
for fp in ['C:/Windows/Fonts/meiryo.ttc', 'C:/Windows/Fonts/YuGothM.ttc', 'C:/Windows/Fonts/msgothic.ttc']:
    if os.path.exists(fp):
        try:
            fm.fontManager.addfont(fp)
            plt.rcParams['font.family'] = fm.FontProperties(fname=fp).get_name()
            break
        except Exception:
            pass

RED = colors.HexColor('#e74c3c'); AMBER = colors.HexColor('#f39c12'); GREEN = colors.HexColor('#27ae60')
NAVY = colors.HexColor('#2c3e6b'); LGRAY = colors.HexColor('#f4f5f7'); MGRAY = colors.HexColor('#dde1e8'); WHITE = colors.white
C_RED = '#e74c3c'; C_GREEN = '#27ae60'; C_NAVY = '#2c3e6b'

# ============================================================
# 既定データ（空の雛形）。init コマンドはこの内容からExcelシートを新規作成する。
# 実際の月次数値・個人名等はここには書かない（公開リポジトリのため）。
# 通常運用は「前月の完成済みxlsxをコピーして今月分にする」ほうが早い。
# ============================================================
DEFAULT_DATA = {
    'report_month': '○月',
    'page_label': '20XX年○月',
    'top_notice': [
        '（重要事項・経営層コメントをここに記入）',
        '【前月フォロー】（前月の指摘に対する今月の状況を記入）',
    ],
    'trend': {
        'months': ['1月', '2月', '3月', '4月', '5月', '6月'],
        'defect': [0, 0, 0, 0, 0, 0],
        'defect_target': 41666,
        'profit': [0, 0, 0, 0, 0, 0],
        'profit_target': 8000,
    },
    'quality': {
        'header': ['指標', '今月実績', '月次目標', '前月比・状況'],
        'rows': [
            ['外部苦情件数', '?件', '1.08件/月', ''],
            ['工程内不良金額', '?円', '41,666円', ''],
            ['社内選別・修正', '?円', '37,916円', ''],
            ['外部業者選別費', '?円', '12,500円', ''],
        ],
        'note': '※品質の補足コメントを記入',
    },
    'mfg': {
        'header': ['項目', '内容・状況', '判定'],
        'rows': [
            ['項目A', '内容を記入', '？'],
            ['項目B', '内容を記入', '？'],
        ],
    },
    'asm': {
        'header': ['項目', '内容・状況'],
        'rows': [
            ['計画（前月対策）', '内容を記入'],
            ['実績', '内容を記入'],
            ['対計画達成率', '内容を記入'],
        ],
    },
    'tech': {
        'header': ['項目', '内容・状況', '判定'],
        'rows': [
            ['新規金型', '内容を記入', '？'],
            ['改修状況', '内容を記入', '？'],
        ],
    },
    'dlv': {
        'header': ['項目', '内容・状況', '判定'],
        'rows': [
            ['納入異常', '?件', '？'],
            ['輸送費', '売上対比?%', '？'],
        ],
    },
    'press': {
        'header': ['指標', '前月確定', '今月速報', '判定'],
        'rows': [
            ['粗利（金型売上除く）', '?千円', '?千円', '？'],
            ['計画達成率', '—', 'Aライン ?%／Bライン ?%／Cライン ?%', '？'],
        ],
        'note': '※プレス実績の補足コメントを記入',
    },
    'follow': {
        'header': ['部署', '前月の指摘内容', '今月の状況', '評価'],
        'rows': [
            ['製造T(見原)', '前月の指摘内容', '今月の状況', '？'],
            ['品管T(八木)', '前月の指摘内容', '今月の状況', '？'],
            ['技術T', '前月の指摘内容', '今月の状況', '？'],
            ['納入管理(木村)', '前月の指摘内容', '今月の状況', '？'],
            ['組立検査(永丘)', '前月の指摘内容', '今月の状況', '？'],
        ],
    },
    'inst': {
        'header': ['対象', '指摘・指示内容', '重要度'],
        'rows': [
            ['全社', '指示内容を記入', '★★★'],
        ],
    },
    'act': {
        'header': ['No.', '内容', '担当', '期限'],
        'rows': [
            ['1', 'アクション内容', '担当者', '期限'],
        ],
    },
}

# ============================================================
# SHEET SCHEMA（Excel入力シート構成）
#  設定        : A列キー / B列値（report_month, page_label, defect_target, profit_target）
#  重要事項     : A列に段落を1行1段落で入力（複数行あれば別パラグラフとして表示）
#  指標推移     : ヘッダ行 + 月別の 工程内不良金額(円) / 粗利(千円)
#  品質実績〜重要アクション: 各テーブルごとに1シート。1行目=見出し、以降=データ行。
#                            末尾に「※注記」行があれば note として扱う（品質実績・プレス実績のみ）
# ============================================================
SHEET_TABLES = [
    ('品質実績', 'quality', True),
    ('製造実績', 'mfg', False),
    ('組立実績', 'asm', False),
    ('技術実績', 'tech', False),
    ('納入管理実績', 'dlv', False),
    ('プレス実績', 'press', True),
    ('前月フォロー', 'follow', False),
    ('来月指示', 'inst', False),
    ('重要アクション', 'act', False),
]

HEADER_FILL = PatternFill('solid', fgColor='2C3E6B')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
NOTE_FONT = Font(color='C0392B', italic=True, size=9)
WRAP = Alignment(wrap_text=True, vertical='top')


def write_template_xlsx(data, path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 設定
    ws = wb.create_sheet('設定')
    ws.append(['キー', '値'])
    for c in ws[1]:
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    ws.append(['report_month', data['report_month']])
    ws.append(['next_month', data.get('next_month', _guess_next_month(data['report_month']))])
    ws.append(['page_label', data['page_label']])
    ws.append(['defect_target_yen', data['trend']['defect_target']])
    ws.append(['profit_target_senyen', data['trend']['profit_target']])
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 40

    # 重要事項
    ws = wb.create_sheet('重要事項')
    ws.append(['段落（1行=1パラグラフ。改行して複数段落にしてOK）'])
    ws['A1'].font = HEADER_FONT; ws['A1'].fill = HEADER_FILL
    for para in data['top_notice']:
        ws.append([para])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
    ws.column_dimensions['A'].width = 120

    # 指標推移
    ws = wb.create_sheet('指標推移')
    ws.append(['月', '工程内不良金額(円)', '粗利(千円)'])
    for c in ws[1]:
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    months = data['trend']['months']
    defect = data['trend']['defect']
    profit = data['trend']['profit']
    for i, m in enumerate(months):
        ws.append([m, defect[i], profit[i]])
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16

    # 各テーブルシート
    for sheet_name, key, has_note in SHEET_TABLES:
        d = data[key]
        ws = wb.create_sheet(sheet_name)
        ws.append(d['header'])
        for c in ws[1]:
            c.font = HEADER_FONT; c.fill = HEADER_FILL
        for row in d['rows']:
            ws.append(row)
        ncols = len(d['header'])
        for i in range(ncols):
            col = get_column_letter(i + 1)
            ws.column_dimensions[col].width = 55 if i == 1 else 18
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = WRAP
        if has_note and d.get('note'):
            r = ws.max_row + 2
            ws.cell(row=r, column=1, value='※注記').font = NOTE_FONT
            ws.cell(row=r, column=2, value=d['note']).font = NOTE_FONT
            ws.cell(row=r, column=2).alignment = WRAP

    wb.save(path)
    print('Excel入力シートを作成しました:', path)


def _cell(v):
    return '' if v is None else str(v).strip()


def _guess_next_month(report_month):
    # "6月" -> "7月"（12月の次は1月）。当月表記の自動推定に使う既定値。
    import re
    m = re.match(r'(\d+)', str(report_month))
    if not m:
        return report_month
    n = int(m.group(1)) % 12 + 1
    return f'{n}月'


def read_data_from_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    data = {}

    ws = wb['設定']
    kv = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2) if row[0].value}
    data['report_month'] = _cell(kv.get('report_month', ''))
    nm = _cell(kv.get('next_month', ''))
    data['next_month'] = nm if nm else _guess_next_month(data['report_month'])
    data['page_label'] = _cell(kv.get('page_label', ''))
    defect_target = int(kv.get('defect_target_yen') or 41666)
    profit_target = int(kv.get('profit_target_senyen') or 8000)

    ws = wb['重要事項']
    top_notice = [_cell(row[0].value) for row in ws.iter_rows(min_row=2) if _cell(row[0].value)]
    data['top_notice'] = top_notice

    ws = wb['指標推移']
    months, defect, profit = [], [], []
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        months.append(_cell(row[0].value))
        defect.append(int(row[1].value or 0))
        profit.append(int(row[2].value or 0))
    data['trend'] = {'months': months, 'defect': defect, 'defect_target': defect_target,
                      'profit': profit, 'profit_target': profit_target}

    for sheet_name, key, has_note in SHEET_TABLES:
        ws = wb[sheet_name]
        rows_all = list(ws.iter_rows(values_only=True))
        header = [ _cell(v) for v in rows_all[0] ]
        body_rows = []
        note = None
        for r in rows_all[1:]:
            if r[0] is not None and _cell(r[0]) == '※注記':
                note = _cell(r[1]) if len(r) > 1 else ''
                continue
            if all(v is None or _cell(v) == '' for v in r):
                continue
            body_rows.append([_cell(v) for v in r[:len(header)]])
        d = {'header': header, 'rows': body_rows}
        if has_note:
            d['note'] = note or ''
        data[key] = d

    return data


# ============================================================
# PDF生成（reportlab）
# ============================================================
def make_styles():
    s = {}
    s['title'] = ParagraphStyle('title', fontSize=13, fontName='HeiseiKakuGo-W5', textColor=colors.HexColor('#1a1a2e'), alignment=TA_CENTER, spaceAfter=2, leading=16)
    s['sec'] = ParagraphStyle('sec', fontSize=9, fontName='HeiseiKakuGo-W5', textColor=WHITE, backColor=NAVY, leftIndent=4, spaceBefore=4, spaceAfter=2, leading=13)
    s['sec2'] = ParagraphStyle('sec2', fontSize=9, fontName='HeiseiKakuGo-W5', textColor=WHITE, backColor=colors.HexColor('#7f1d1d'), leftIndent=4, spaceBefore=4, spaceAfter=2, leading=13)
    s['note'] = ParagraphStyle('note', fontSize=7, fontName='HeiseiKakuGo-W5', textColor=colors.HexColor('#c0392b'), leading=10, leftIndent=4)
    s['label'] = ParagraphStyle('label', fontSize=6.5, fontName='HeiseiKakuGo-W5', textColor=colors.HexColor('#555555'), alignment=TA_CENTER, leading=9)
    s['notice'] = ParagraphStyle('notice', fontSize=7.3, fontName='HeiseiKakuGo-W5', textColor=colors.HexColor('#1a1a2e'), leading=11.5, leftIndent=4, rightIndent=4,
                                  backColor=colors.HexColor('#fef9e7'), borderColor=colors.HexColor('#e0c66b'), borderWidth=0.6, borderPadding=4, spaceAfter=4)
    s['cell'] = ParagraphStyle('cell', fontSize=6.6, fontName='HeiseiKakuGo-W5', textColor=colors.HexColor('#222222'), leading=9.2)
    s['cellc'] = ParagraphStyle('cellc', fontSize=6.8, fontName='HeiseiKakuGo-W5', textColor=colors.HexColor('#222222'), leading=9.2, alignment=TA_CENTER)
    s['cellc_ok'] = ParagraphStyle('cellc_ok', parent=s['cellc'], textColor=GREEN)
    s['cellc_warn'] = ParagraphStyle('cellc_warn', parent=s['cellc'], textColor=AMBER)
    s['cellc_bad'] = ParagraphStyle('cellc_bad', parent=s['cellc'], textColor=RED)
    return s


def tbl_style_base():
    return [('FONTNAME', (0, 0), (-1, -1), 'HeiseiKakuGo-W5'), ('FONTSIZE', (0, 0), (-1, -1), 6.8),
            ('BACKGROUND', (0, 0), (-1, 0), NAVY), ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [LGRAY, WHITE]), ('GRID', (0, 0), (-1, -1), 0.4, MGRAY),
            ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 3), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]


_JUDGE_STYLE = {'○': 'cellc_ok', '△': 'cellc_warn', '×': 'cellc_bad'}


def make_table(S, header, rows, col_widths, extra_style=None):
    wrapped = [header]
    for row in rows:
        wrapped.append([
            Paragraph(str(v).replace('\n', '<br/>'), S[_JUDGE_STYLE.get(str(v), 'cellc' if len(str(v)) <= 6 else 'cell')])
            for v in row
        ])
    style = tbl_style_base()
    if extra_style:
        style += extra_style
    t = Table(wrapped, colWidths=col_widths)
    t.setStyle(TableStyle(style))
    return t


def chart_defect_trend(trend):
    months, defect, target = trend['months'], trend['defect'], trend['defect_target']
    fig, ax = plt.subplots(figsize=(5.5, 2.6))
    bcols = [C_RED if v > target else C_GREEN for v in defect]
    bars = ax.bar(months, [v / 1000 for v in defect], color=bcols, alpha=0.85, width=0.55, zorder=3)
    ax.axhline(target / 1000, color=C_NAVY, lw=1.3, ls='--', label=f'月次目標{target/1000:.1f}千円')
    ax.set_ylabel('千円', fontsize=7); ax.tick_params(axis='both', labelsize=7)
    ax.yaxis.grid(True, linestyle='--', alpha=0.4, zorder=0); ax.set_axisbelow(True)
    for bar, v in zip(bars, defect):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 3, f'{v/1000:.0f}', ha='center', fontsize=6, color='#333')
    ax.legend(fontsize=7, framealpha=0.8); ax.set_title(f'工程内不良金額推移（{months[0]}〜{months[-1]}）', fontsize=8.5, fontweight='bold', pad=4)
    fig.tight_layout(pad=0.5); buf = io.BytesIO(); fig.savefig(buf, format='png', dpi=150, bbox_inches='tight'); plt.close(fig); buf.seek(0); return buf


def chart_profit_trend(trend):
    months, profit, target = trend['months'], trend['profit'], trend['profit_target']
    fig, ax = plt.subplots(figsize=(5.5, 2.6))
    bcols = [C_GREEN if v >= 0 else C_RED for v in profit]
    bars = ax.bar(months, profit, color=bcols, alpha=0.85, zorder=3, width=0.55)
    ax.axhline(target, color='#333', lw=1.3, ls='--', label=f'目標{target:,}千円')
    ax.axhline(0, color='#333', lw=0.7)
    ax.set_ylabel('千円', fontsize=7); ax.tick_params(axis='both', labelsize=7)
    ax.yaxis.grid(True, linestyle='--', alpha=0.4, zorder=0); ax.set_axisbelow(True)
    for bar, v in zip(bars, profit):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 150 if v >= 0 else bar.get_height() - 500, f'{v:+,}', ha='center', fontsize=6, color='#333')
    ax.legend(fontsize=7, framealpha=0.8); ax.set_title(f'月次粗利推移（{months[0]}〜{months[-1]}・金型売上除く）', fontsize=8.5, fontweight='bold', pad=4)
    fig.tight_layout(pad=0.5); buf = io.BytesIO(); fig.savefig(buf, format='png', dpi=150, bbox_inches='tight'); plt.close(fig); buf.seek(0); return buf


def build_pdf(data, out_path):
    S = make_styles()
    doc = SimpleDocTemplate(out_path, pagesize=A4, leftMargin=10 * mm, rightMargin=10 * mm, topMargin=6 * mm, bottomMargin=5 * mm)
    story = []
    month = data['report_month']; page_label = data['page_label']

    story.append(Paragraph(f'71期 月次会　{month}実績', S['title']))
    story.append(HRFlowable(width='100%', thickness=1.5, color=NAVY, spaceAfter=3))
    story.append(Paragraph('■ 重要事項・前月フォロー', S['sec2']))
    for para in data['top_notice']:
        story.append(Paragraph(para.replace('\n', '<br/>'), S['notice']))

    q = data['quality']
    story.append(Paragraph(f'■ {month} 品質実績（八木）', S['sec']))
    story.append(make_table(S, q['header'], q['rows'], [46 * mm, 26 * mm, 26 * mm, 82 * mm]))
    if q.get('note'):
        story.append(Paragraph(q['note'], S['note']))

    m = data['mfg']
    story.append(Paragraph(f'■ {month} 製造実績（見原）', S['sec']))
    story.append(make_table(S, m['header'], m['rows'], [34 * mm, 132 * mm, 14 * mm], [('VALIGN', (0, 0), (-1, -1), 'TOP')]))

    a = data['asm']
    story.append(Paragraph(f'■ {month} 組立実績（永丘）', S['sec']))
    story.append(make_table(S, a['header'], a['rows'], [30 * mm, 150 * mm], [('VALIGN', (0, 0), (-1, -1), 'TOP')]))

    t = data['tech']
    story.append(Paragraph(f'■ {month} 技術実績（大石・遠藤）', S['sec']))
    story.append(make_table(S, t['header'], t['rows'], [30 * mm, 136 * mm, 14 * mm], [('VALIGN', (0, 0), (-1, -1), 'TOP')]))

    dl = data['dlv']
    story.append(Paragraph(f'■ {month} 納入管理実績（木村）', S['sec']))
    story.append(make_table(S, dl['header'], dl['rows'], [24 * mm, 142 * mm, 14 * mm], [('VALIGN', (0, 0), (-1, -1), 'TOP')]))

    story.append(Spacer(1, 2 * mm)); story.append(HRFlowable(width='100%', thickness=0.8, color=MGRAY))
    story.append(Paragraph(f'橋本工業株式会社　71期 月次会資料　{page_label}　（1/2）', S['label']))

    story.append(PageBreak())
    story.append(Paragraph('71期 月次会　― プレス実績・前月フォロー・指示事項 ―', S['title']))
    story.append(HRFlowable(width='100%', thickness=1.5, color=NAVY, spaceAfter=3))

    p = data['press']
    story.append(Paragraph(f'■ {month} プレス実績（深津英介）', S['sec']))
    story.append(make_table(S, p['header'], p['rows'], [42 * mm, 54 * mm, 60 * mm, 14 * mm], [('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    if p.get('note'):
        story.append(Paragraph(p['note'], S['note']))

    story.append(Paragraph('■ 主要指標推移', S['sec']))
    img1 = Image(chart_defect_trend(data['trend']), width=88 * mm, height=38 * mm)
    img2 = Image(chart_profit_trend(data['trend']), width=88 * mm, height=38 * mm)
    ct = Table([[img1, img2]], colWidths=[90 * mm, 90 * mm])
    ct.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                             ('LEFTPADDING', (0, 0), (-1, -1), 1), ('RIGHTPADDING', (0, 0), (-1, -1), 1),
                             ('TOPPADDING', (0, 0), (-1, -1), 1), ('BOTTOMPADDING', (0, 0), (-1, -1), 0)]))
    story.append(ct)

    fo = data['follow']
    story.append(Paragraph('■ 前月実績報告 指摘事項フォローアップ', S['sec2']))
    story.append(make_table(S, fo['header'], fo['rows'], [20 * mm, 52 * mm, 88 * mm, 18 * mm], [('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    story.append(Paragraph('○:改善　△:継続課題　×:未対応', S['note']))

    next_month = data.get('next_month') or _guess_next_month(month)
    ins = data['inst']
    story.append(Paragraph(f'■ {next_month} 指摘・指示事項', S['sec']))
    story.append(make_table(S, ins['header'], ins['rows'], [26 * mm, 143 * mm, 11 * mm], [('VALIGN', (0, 0), (-1, -1), 'TOP')]))

    act = data['act']
    story.append(Paragraph(f'■ {next_month} 重要対応アクション', S['sec']))
    story.append(make_table(S, act['header'], act['rows'], [8 * mm, 114 * mm, 32 * mm, 26 * mm], [('VALIGN', (0, 0), (-1, -1), 'TOP')]))

    story.append(Spacer(1, 1 * mm)); story.append(HRFlowable(width='100%', thickness=0.8, color=MGRAY))
    story.append(Paragraph(f'橋本工業株式会社　71期 月次会資料　{page_label}　（2/2）', S['label']))

    doc.build(story)
    print('PDFを作成しました:', out_path)


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    cmd = sys.argv[1]
    if cmd == 'init':
        write_template_xlsx(DEFAULT_DATA, sys.argv[2])
    elif cmd == 'build':
        xlsx_path = sys.argv[2]
        pdf_path = sys.argv[3] if len(sys.argv) > 3 else os.path.splitext(xlsx_path)[0] + '.pdf'
        data = read_data_from_xlsx(xlsx_path)
        build_pdf(data, pdf_path)
    else:
        print(__doc__)
        sys.exit(1)
